# Spreadsheet plugin class
import re
import os
import ast
from openpyxl import load_workbook
import inspect
import json

from edatasheets_creator.document.jsondatasheet import JsonDataSheet
from edatasheets_creator.document.jsondatasheetschema import JsonDataSheetSchema
from edatasheets_creator.utility.path_utilities import validateRealPath
from edatasheets_creator.utility.format import Format
from edatasheets_creator.utility.filevalidation import readWorkgroupSchema, validateWithSchema, writeToJSON


from edatasheets_creator.functions import t
from collections import deque
from edatasheets_creator.constants import datasheetconstants
from edatasheets_creator.constants import serializationconstants
from edatasheets_creator.constants import common_constants
from edatasheets_creator.constants import schema_constants
from edatasheets_creator.constants import spreadsheettypes
from edatasheets_creator.exceptions.configurationerror import ConfigurationError
from edatasheets_creator.plugins.spreadsheetmap import SpreadsheetMap
from edatasheets_creator.logger.exceptionlogger import ExceptionLogger


class Plugin:
    """
    Spreadsheet plugin class that implements datasheet generation from an XLSX
    """

    def __init__(self):
        """
        Class initialization
        """

        self._worksheetSectionIndexWritten = False  # used in multi section worksheets to indicate that index values are already written
        self._indexOnRow = -1
        self._indexOnCol = -1
        self.format = Format()
        self.startIndexCurrentTable = 0
        self.tableCounter = 1

    def __repr__(self):
        """
            Returns a name for the class

        Returns:
            string : class name
        """
        return __name__ + '.' + inspect.currentframe().f_code.co_name

    def process(self, inputFileName, outputFileName, mapFileName=""):
        """

        Main logic method for spreadsheet processing

        Args:
            inputFileName (PosixPath): Input file name
            outputFileName (PosixPath): Output file name
            mapFileName (PosixPath): Map file to guide parser
        """

        datasheet = {}

        try:
            msg = t("Spreadsheet Plugin is loaded")
            ExceptionLogger.logInformation(__name__, msg)

            # Validate if the input files exists on the system as they are required
            if (not validateRealPath(inputFileName)):
                # Input file does not exist
                ExceptionLogger.logInformation(__name__, "", t("\n Input file does not exists"))
                print()
                return

            if ((not validateRealPath(mapFileName)) and mapFileName != ""):
                # Map file does not exist
                ExceptionLogger.logInformation(__name__, "", t("\n Map file does not exists"))
                print()
                return

            self._fileName = inputFileName
            self._outputFileName = outputFileName
            self._mapFileName = mapFileName

            wb = load_workbook(filename=inputFileName, data_only=True)
            ExceptionLogger.logInformation(__name__, t("\n\nProcessing") + " " + t("workbook") + ":  " + str(inputFileName) + "...\n")

            outputSuffix = os.path.splitext(self._outputFileName)
            if outputSuffix[1] == ("." + serializationconstants.C_HEADER_NAME):
                msg1 = t("\nCreating Header File\n")
                ExceptionLogger.logInformation(__name__, msg1)
                # Header.createHeaderFile(wb,self._fileName,self._outputFileName)
                self.createHeaderFile(wb, inputFileName, outputFileName)

            if self._mapFileName == "":
                msg1 = t("\nNo map file so will use default processing\n")
                ExceptionLogger.logInformation(__name__, msg1)
                self.processWithoutMap(wb, inputFileName, outputFileName, datasheet)
            if self._mapFileName != "":

                # Load map file
                map = SpreadsheetMap(self._mapFileName)

                # get the sheet names
                sheetNames = map.getSheetNames()

                # get the title
                baseName = os.path.basename(inputFileName).split('.')[0]
                datasheetTitle = datasheetconstants.DATASHEET_DEFAULT_TITLE + ' - ' + baseName

                # get the datasheet description
                datasheetDescription = self.getDatasheetDescription(wb, map.getDatasheetDescriptionLocation())

                if datasheetDescription is not None:
                    datasheetDescription = datasheetDescription.replace('\n', ' ')

                jds = JsonDataSheet(self._outputFileName)  # outputfile

                # generate the datasheet header
                datasheetHeader = jds.setMetadata(datasheetTitle, datasheetDescription, self._fileName)

                # add header information to the datasheet if metadata inclusion in map file
                if map.includeMetadata():
                    datasheet = datasheetHeader

                # ExceptionLogger.logDebug(__name__,"datasheet:",datasheet)

                # iterate through the sheets described in the map file, ignore other worksheets

                datasheet[spreadsheettypes.SPREADSHEET_TABLES] = {}
                for i in sheetNames:
                    self.startIndexCurrentTable = 0
                    if map.ignoreBlanks(i) is False:
                        ExceptionLogger.logInformation(__name__, t("\nProcessing") + " " + t("spreadsheet") + ":  " + i + "..." + " Output for " + i + " may not match schema. Set ignoreBlanks in map file to True to correct")
                    else:
                        ExceptionLogger.logInformation(__name__, t("\nProcessing") + " " + t("spreadsheet") + ":  " + i + "...")
                    self._worksheetSectionIndexWritten = False  # controls writing the IndexOn item
                    ws = wb[i]

                    for i, s in enumerate(wb.sheetnames):
                        if s == ws.title:
                            break

                    wb.active = i
                    sheet = wb.active

                    # if the worksheet should be included in the datasheet, process it
                    if map.includeInDatasheet(sheet.title):
                        self.tableCounter = 1
                        self._indexOnCol = -1
                        self._indexOnRow = -1
                        self.parseWorksheet(datasheet, wb, sheet.title, map)

                        # ExceptionLogger.logDebug(__name__,"datasheet:",v)
                if map.checkIndustryFormat():
                    datasheet = self.applySpecialRules(datasheet, map, sheetNames)
                    strMsg = "\n\n" + t("Validating ") + " " + str((self._outputFileName)) + " with schema " + "...\n"
                    ExceptionLogger.logInformation(__name__, strMsg)
                    result, errorMsg = validateWithSchema(datasheet[spreadsheettypes.SPREADSHEET_TABLES], map.getComponentType())

                    if result is True:
                        datasheet = datasheet[spreadsheettypes.SPREADSHEET_TABLES]
                        edatasheet = datasheet

                        writeToJSON(edatasheet, self._outputFileName)

                        # write the schema
                        schema = JsonDataSheetSchema(self._outputFileName)
                        schema.write()
                    else:
                        ExceptionLogger.logInformation(__name__, errorMsg)

                else:
                    edatasheet = {spreadsheettypes.SPREADSHEET_DATASHEET: datasheet}
                    strMsg = "\n\n" + t("Writing") + " " + str((self._outputFileName)) + "...\n"
                    ExceptionLogger.logInformation(__name__, strMsg)

                    # pretty print JSON, preserving unicode characters
                    writeToJSON(edatasheet, self._outputFileName)

                    # write the schema
                    schema = JsonDataSheetSchema(self._outputFileName)
                    schema.write()

        except FileNotFoundError as fnf:
            ExceptionLogger.logError(__name__, "", fnf)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def processWithoutMap(self, wb, inputFileName, outputFileName, datasheet):
        """

       Main logic method for spreadsheet processing without a map file needed

       Args:
           wb : Workbook of the given file name
           inputFileName (PosixPath): Input file name
           outputFileName (PosixPath): Output file name
           datasheet : The outputted data
       """
        try:
            # defining all the needed variables
            self._wb = wb
            self._inputFileName = inputFileName
            self._outputFileName = outputFileName
            self._datasheet = datasheet

            indexOnDict = dict()

            # get the title
            baseName = os.path.basename(inputFileName).split('.')[0]
            datasheetTitle = datasheetconstants.DATASHEET_DEFAULT_TITLE + ' - ' + baseName
            dataSheetDescription = " "
            jds = JsonDataSheet(self._outputFileName)  # outputfile
            datasheetHeader = jds.setMetadata(datasheetTitle, dataSheetDescription, self._inputFileName)

            # put the datasheet header on the datasheet
            self._datasheet = datasheetHeader

            self._datasheet[spreadsheettypes.SPREADSHEET_TABLES] = {}
            self._tablesDatasheet = self._datasheet[spreadsheettypes.SPREADSHEET_TABLES]

            # iterate through the sheets of the excel file
            for i in wb.sheetnames:

                worksheetName = JsonDataSheet.generateValidJsonFieldName(str(i))
                self._tablesDatasheet[worksheetName] = []

                ExceptionLogger.logInformation(__name__, t("\nProcessing") + " " + t("spreadsheet") + ":  " + i + "...")

                self._worksheetSectionIndexWritten = False  # controls writing the IndexOn item
                self._indexOnCol = 0
                self._indexOnRow = 1
                self._colLetter = SpreadsheetMap.getColumnLetter(self._indexOnCol)

                cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)
                indexFieldName = cellContents
                values = []

                # while we are not at the end of the excel spreadsheet, keep iterating through it
                while self._indexOnRow != (wb[i].max_row) and self._indexOnCol != (wb[i].max_column):
                    self._indexOnRow += 1

                    if self._indexOnRow == 1:
                        cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)
                        indexFieldName = cellContents
                    else:
                        cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)
                        values.append(cellContents)

                    if self._indexOnRow == wb[i].max_row and self._indexOnCol != wb[i].max_column:
                        indexOnDict[indexFieldName] = values
                        values = []
                        self._indexOnRow = 0
                        self._indexOnCol += 1
                    self._colLetter = SpreadsheetMap.getColumnLetter(self._indexOnCol)

                # adding the values of the dictionary to the output file
                strMsg = "\n\n" + t("Writing") + " " + str((self._outputFileName)) + "...\n"
                ExceptionLogger.logInformation(__name__, strMsg)
                self._tablesDatasheet[worksheetName].append(indexOnDict)
                indexOnDict = dict()

                edatasheet = {spreadsheettypes.SPREADSHEET_DATASHEET: self._datasheet}

                # pretty print JSON, preserving unicode characters
                with open(self._outputFileName, "w", encoding='utf-8') as outfile:
                    json.dump(edatasheet, outfile, ensure_ascii=False, indent=serializationconstants.JSON_INDENT_DEFAULT)
                    outfile.close()

                # writing the schema
                schema = JsonDataSheetSchema(self._outputFileName)
                schema.write()

        except FileNotFoundError as fnf:
            ExceptionLogger.logError(__name__, "", fnf)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def createHeaderFile(self, wb, inputFileName, outputFileName):
        """

       Main logic method for creating a header file given user inputted a .h file

       Args:
           wb : Workbook of the given file name
           inputFileName (PosixPath): Input file name
           outputFileName (PosixPath): Output file name
       """
        try:

            self._inputFileName = inputFileName
            self._wb = wb
            self._outputFileName = outputFileName
            with open(outputFileName, "a") as outputFile:

                for i in wb.sheetnames:
                    self._indexOnCol = 0
                    self._indexOnRow = 2
                    self._colLetter = SpreadsheetMap.getColumnLetter(self._indexOnCol)

                    cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)

                    while self._indexOnRow != (wb[i].max_row) and self._indexOnCol != (wb[i].max_column):
                        self._indexOnCol += 1

                        if self._indexOnCol == (wb[i].max_column):
                            cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)
                            outputFile.write(cellContents + "\n")
                        else:
                            cellContents = self.getCellValue(wb, i, self._indexOnRow, self._colLetter)
                            outputFile.write("#define ")
                            outputFile.write(cellContents + "   ")

                        if self._indexOnRow != wb[i].max_row and self._indexOnCol == wb[i].max_column:
                            self._indexOnCol = 0
                            self._indexOnRow += 1
                        self._colLetter = SpreadsheetMap.getColumnLetter(self._indexOnCol)

        except FileNotFoundError as fnf:
            ExceptionLogger.logError(__name__, "", fnf)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getDatasheetDescription(self, wb, descriptionLocation):
        """
        Returns a datasheet description based referenced from the specified location in the workbook.  This can be in any worksheet.

        Args:
            wb (workbook): Spreadsheet workbook object
            descriptionLocation (dict): Contains the worksheet name, row and column containing the description

        Returns:
            string: description
        """

        s = datasheetconstants.DATASHEET_DEFAULT_DESCRIPTION  # initialize to default

        try:
            if (len(descriptionLocation) > 0):

                # get the location information for the datasheet description
                sheetName = descriptionLocation[spreadsheettypes.SPREADSHEET_MAP_SHEETNAME_FIELD]
                rowNum = descriptionLocation[spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                colLetter = descriptionLocation[spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]

                # set the active worksheet to the sheet containing the datasheet description
                # ws = wb[sheetName]
                wb.active = wb.worksheets.index(wb[sheetName])
                sheet = wb.active

                # build the cell name
                cell = colLetter + str(rowNum)

                # get the value from the cell
                s = sheet[cell].value

                return s

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getCellValue(self, wb, sheetName, row, col):
        """

        Returns the value contained in a cell within a workbook and worksheet.

        Args:
            wb (openpyxl workbook): The workbook to reference
            sheetName (string): The worksheet within the workbook to reference
            row (number): The row to reference
            col (number): The cell to reference

        Returns:
            Any : The value contained within the cell.
        """
        try:
            # set the active worksheet to the sheet containing the datasheet description
            # ws = wb[sheetName]
            wb.active = wb.worksheets.index(wb[sheetName])
            sheet = wb.active

            # build the cell name
            cell = col + str(row)

            # get the value from the cell
            if (self.isMerged(wb, row, col)):
                # Row num plus one to make it match with the excel numeration
                rng = [s for s in sheet.merged_cells.ranges if sheet[cell].coordinate in s]
                return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value
            elif sheet[cell].value is None:
                # cellValue = "defaultValue"
                return "empty"
            return sheet[cell].value
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getColValueFromRow(self, rowObj, col, wb=None, rowNum=None):
        """
        Returns the value contained within the specified column from a row object.  This method should be used when an entire row of data is read
        from the workbook.

        Returns:
            Any: The value contained within the cell.
        """
        try:

            if rowObj is not None and col is not None:
                colIdx = SpreadsheetMap.getColumnIndex(col)
                cell = rowObj[colIdx]
                cellValue = cell.value
                if cellValue is not None and isinstance(cellValue, str):
                    cellValue = cellValue.replace("_x000D_", "")
                if (cellValue is None):
                    if (wb is None or rowNum is None):
                        msg = "There was a problem retrieving the column value, the value is null."
                        ExceptionLogger.logError(__name__, msg)
                        defaultValue = "defaultValue"
                        return defaultValue
                    else:
                        if (self.isMerged(wb, rowNum + 1, col)):
                            # Row num plus one to make it match with the excel numeration
                            cellValue = self.getCellValueIfMerged(wb, rowNum + 1, col)
                        else:
                            cellValue = " "

                # apply format to the cell value if it is a string
                cellType = type(cellValue)
                if cellType == str:
                    cellValue = self.format.format_value(cellValue)
                elif str(cellType) == spreadsheettypes.SPREADSHEET_DATETIME_TYPE:
                    cellValue = str(cellValue)
                return cellValue

            else:
                msg = "There was a problem retrieving the column value from the row object"
                ExceptionLogger.logError(__name__, msg)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getCellValueIfMerged(self, wb, row, col):
        try:
            returnValue = None
            currentCell = str(col) + str(row)
            # Col plus 1 because the bounds in the merged ranges start in 1
            colIndex = SpreadsheetMap.getColumnIndex(currentCell) + 1
            sheet = wb.active

            for range_ in sheet.merged_cells.ranges:
                min_row = range_.min_row
                min_col = range_.min_col
                max_row = range_.max_row
                max_col = range_.max_col
                if (min_row <= row and min_col <= colIndex and row <= max_row and colIndex <= max_col):
                    cells = str(range_).split(':')[0]
                    # ExceptionLogger.logDebug(__name__, "cells=", cells)
                    mergedRowNum = int(re.search("\d+", str(cells))[0])  # noqa
                    # ExceptionLogger.logDebug(__name__, "mergedRowNum=", mergedRowNum)
                    cIdx = SpreadsheetMap.getColumnIndex(cells)
                    colLetter = SpreadsheetMap.getColumnLetter(cIdx)
                    # ExceptionLogger.logDebug(__name__, "cells.value=", self.getCellValue(wb, sheet.title, mergedRowNum, colLetter))
                    returnValue = self.getCellValue(wb, sheet.title, mergedRowNum, colLetter)
                    if returnValue is None:
                        returnValue = ' '
                    break
            return returnValue
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def isMerged(self, wb, row, col):
        try:
            sheet = wb.active
            cell = sheet.cell(row, SpreadsheetMap.getColumnIndex(col) + 1)
            # strMsg = "row=" + str(row) + ", col=" + str(col)
            # ExceptionLogger.logDebug(__name__, strMsg)

            # ExceptionLogger.logDebug(__name__, "merged_cells.ranges=", sheet.merged_cells.ranges)
            for m in sheet.merged_cells.ranges:

                # ExceptionLogger.logDebug(__name__, "m=", m)
                # ExceptionLogger.logDebug(__name__, "cell.coordinate=", cell.coordinate)
                if (cell.coordinate in m):
                    # ExceptionLogger.logDebug(__name__, "Yes, it's merged!", m)
                    return True
            return False

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getRow(self, wb, sheetName, row):
        """
        Reads and returns the specified row from a worksheet within a workbook

        Args:
            wb (workbook): The workbook to reference
            sheetName (string): The worksheet to reference
            row (number): The row to read and return

        Returns:
            Row Object: A row object containing the cells and values from the worksheet.
        """
        try:
            # set the active worksheet to the sheet containing the datasheet description
            # ws = wb[sheetName]
            wb.active = wb.worksheets.index(wb[sheetName])
            sheet = wb.active

            # ExceptionLogger.logDebug(__name__,"Merged cells=",sheet.merged_cells)
            # ExceptionLogger.logDebug(__name__,"Merged cellRanges=",sheet.merged_cell_ranges)
            return sheet[row + 1]

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def parseWorksheet(self, datasheet, wb, sheetName, map):
        """
        Main worksheet parser entry method.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            wb (workbook): The workbook object containing the worksheet.
            sheetName (string): The worksheet name contained in the workbook to parse.
            map (dict): A map containing the parser rules for the worksheet
        """

        try:

            # ExceptionLogger.logDebug(__name__,"datasheet inside this func",datasheet)

            # activate the worksheet
            # ws = wb[sheetName]
            wb.active = wb.worksheets.index(wb[sheetName])
            # sheet = wb.active
            sheetKey = JsonDataSheet.generateValidJsonFieldName(sheetName)

            tablesDatasheet = datasheet[spreadsheettypes.SPREADSHEET_TABLES]
            # tablesDatasheet[sheetKey] = []

            # at this point the active sheet should be set to the worksheet named in the sheetName argument

            sections = map.getSections(sheetName)
            self.parseSections(tablesDatasheet, sheetKey, wb, sheetName, sections, map)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def processFunctionProperties(self, pin_list, macro_state, ignore_blanks):

        """

        This method processes the function properties in pins.

        Args:
            pin_list (list): A list of pins dictionaries with same terminal identifier number which are to be merged to one dictionary
            macro_state (boolean): Flag for if the Excel file is a macro enabled file or not
            ignore_blanks (boolean): Flag for if blanks in the Excel file should be ignored or not
        """

        d = {}
        for k in pin_list[0].keys():
            # store all values asides the function properties in the dictionary
            if k.lower() != datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD.lower():
                d[k] = pin_list[0][k]
            else:
                # merge all function property values into a list and add to dictionary
                function_list = []
                for i in range(len(pin_list)):
                    function_list.append(pin_list[i][datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD][0])
                d[datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD] = function_list
        if macro_state is True:
            d = self.processUnits(d, ignore_blanks)
        return d

    def convertUnitStrToDict(self, unit_string, ignore_blanks, processorType='unit'):
        """

        This method processes the string from a macro file into a dictionary for parsing.

        Args:
            unit_string (string): String consisting of units and their parameters
            ignore_blanks (boolean): Flag for if blanks in the Excel file should be ignored or not
        """
        start_index = 0
        value_start = 0
        unit_list = []
        unit_dictionary = {}
        key = ''
        for i in range(len(unit_string)):
            char = unit_string[i]
            if char == ":" or unit_string[i: i + 10] == datasheetconstants.DATASHEET_CONDITIONS.capitalize():
                if char == ":":
                    if key != datasheetconstants.DATASHEET_CONDITIONS:
                        key = self.format.unit_camel_case(unit_string[start_index:i])
                        start_index = i
                        value_start = i + 1
                else:
                    key = self.format.unit_camel_case(unit_string[start_index:i + 10])
                    start_index = i + 10
                    value_start = i + 11
            if char == ",":
                value_start = i
                value = unit_string[start_index + 1:value_start]
                if (key == datasheetconstants.DATASHEET_CONDITIONS and (unit_string[value_start + 1:value_start + 3].strip() == 'T' or value_start + 1 == len(unit_string))) or (key != datasheetconstants.DATASHEET_CONDITIONS):
                    if self.format.is_float(value.replace(" ", "")):
                        value = float(value.replace(" ", ""))
                        unit_dictionary[key] = value
                    elif len(value) == 0:
                        if ignore_blanks is False:
                            value = ' '
                            unit_dictionary[key] = value
                        else:
                            pass
                    elif isinstance(value, str) and (value.lower() == 'false' or value.lower() == 'true'):
                        unit_dictionary[key] = ast.literal_eval(value.capitalize())
                    elif value.isspace():
                        pass
                    elif isinstance(value, str):
                        unit_dictionary[key] = value.strip()
                    start_index = i + 1
            if processorType == 'unit':
                if unit_string[i: i + 9] == datasheetconstants.DATASHEET_TYP_VALUE and i != 0:
                    unit_list.append(unit_dictionary)
                    unit_dictionary = {}
                    key = ''
            elif processorType == datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS:
                if unit_string[i: i + 14] == 'Component Type' and i != 0:
                    unit_list.append(unit_dictionary)
                    unit_dictionary = {}
        unit_list.append(unit_dictionary)
        return unit_list

    def processUnits(self, sheet_dictionary, ignore_blanks):
        """

        This method processes all units in the macro enabled file.

        Args:
            sheet_dictionary (dict): dictionary of all unit values
            ignore_blanks (boolean): Flag for if blanks in the Excel file should be ignored or not

        Returns:
            sheet_dictionary: Dictionary with units processed
        """
        to_delete_key = []
        for key, value in sheet_dictionary.items():
            if isinstance(value, str):
                if value.startswith(datasheetconstants.DATASHEET_TYP_VALUE):
                    value = value + ","
                    unit_dict = self.convertUnitStrToDict(value, ignore_blanks)
                    if len(unit_dict) == 0:
                        to_delete_key.append(key)
                    sheet_dictionary[key] = unit_dict
        for key in to_delete_key:
            sheet_dictionary.pop(key, None)
        return sheet_dictionary

    def processConnectorWithMacro(self, datasheet, sheetName, macro_state, ignore_blanks):
        """

        This method processes the datasheet for hardware connectors that are macro enabled.

        Args:
            datasheet (dict): A dictionary referencing the datasheet from the initial processing
            sheetName (string):  The name of the worksheet being processed

        """
        refactoredDatasheet = []
        pins = datasheet[sheetName][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]
        start = 0
        temp_start = 0
        while start < len(pins):
            # convert terminal identifier value to string
            if datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD in pins[start]:
                pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD] = str(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD])

            # check for end of pin list
            if start + 1 == len(pins):
                start_list = []
                # check for presence of function properties
                if datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD in pins[start]:
                    start_list.append(pins[start][datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD])
                    pins[start][datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD] = start_list
                if macro_state is True:
                    pins[start] = self.processUnits(pins[start], ignore_blanks)
                refactoredDatasheet.append(pins[start])
                start = start + 1
                temp_start = start

            # check if the next terminal identifier is same as the present one
            elif datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD in pins[start + 1]:
                if temp_start != start or int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) == int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                    try:
                        # count how many terminal identifiers are same
                        while int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) == int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                            start = start + 1
                        # combine all dictionaries into one
                        pins_to_combine = pins[temp_start:start + 1]
                        d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                        refactoredDatasheet.append(d)
                        start = start + 1
                        temp_start = start
                    except Exception:
                        # edge case if pin list count is beyond the number of pins
                        pins_to_combine = pins[temp_start:start + 1]
                        d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                        refactoredDatasheet.append(d)
                        start = start + 1
                        temp_start = start
                # check if the next terminal identifier is different from the present one
                elif int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) != int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                    # check for presence of function properties
                    start_list = []
                    if datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD in pins[start]:
                        start_list.append(pins[start][datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD])
                        pins[start][datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD] = start_list
                    if macro_state is True:
                        pins[start] = self.processUnits(pins[start], ignore_blanks)
                    refactoredDatasheet.append(pins[start])
                    start = start + 1
                    temp_start = start
            else:
                start = start + 1
        return refactoredDatasheet

    def processMacro(self, datasheet, sheetName):
        """

        This method processes the macros in datasheets.

        Args:
            datasheet (dict): A dictionary referencing the datasheet from the initial processing
            sheetName (string):  The name of the worksheet being processed

        """

        refactoredDatasheet = []
        properties = []
        for _, value in datasheet[sheetName].items():
            properties.extend(value)
        start = 0
        while start < len(properties):
            if len(properties[start]) != 0:
                row_value = properties[start]
                d = self.processUnits(row_value, True)
                refactoredDatasheet.append(d)
            else:
                refactoredDatasheet.append(properties[start])
            start = start + 1
        return refactoredDatasheet

    def processConnector(self, datasheet, sheetName, macro_state, ignore_blanks):
        """

        This method processes the datasheet for hardware connectors.

        Args:
            datasheet (dict): A dictionary referencing the datasheet from the initial processing
            sheetName (string):  The name of the worksheet being processed
            macro_state (boolean): Flag for if the Excel file is a macro enabled file or not
            ignore_blanks (boolean): Flag for if blanks in the Excel file should be ignored or not

        Returns:
            refactoredDatasheet: Datasheet with connectors processed
        """

        refactoredDatasheet = []
        try:
            pins = datasheet[sheetName][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]
        except KeyError:
            pins = datasheet[sheetName][datasheetconstants.DATASHEET_PINS_KEYWORD]
        start = 0
        temp_start = 0
        while start < len(pins):
            # convert terminal identifier value to string
            if datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD in pins[start]:
                pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD] = str(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD])
            # check that terminal identifier is not empty since this is the basis for combination into a list
            if len(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) > 0:
                # check for end of pin list
                if start + 1 == len(pins):
                    refactoredDatasheet.append(pins[start])
                    start = start + 1
                    temp_start = start

            # check if the next terminal identifier is same as the present one
                elif datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD in pins[start + 1]:
                    if pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD].isdigit() and pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD].isdigit():
                        if temp_start != start or int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) == int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                            try:
                                # count how many terminal identifiers are same
                                while int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) == int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                                    start = start + 1
                                # combine all dictionaries into one
                                pins_to_combine = pins[temp_start:start + 1]
                                d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                                refactoredDatasheet.append(d)
                                start = start + 1
                                temp_start = start
                            except Exception:
                                # edge case if pin list count is beyond the number of pins
                                pins_to_combine = pins[temp_start:start + 1]
                                d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                                refactoredDatasheet.append(d)
                                start = start + 1
                                temp_start = start
                        # check if the next terminal identifier is different from the present one
                        elif int(pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]) != int(pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]):
                            refactoredDatasheet.append(pins[start])
                            start = start + 1
                            temp_start = start
                    else:
                        if temp_start != start or pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD] == pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]:
                            try:
                                # count how many terminal identifiers are same
                                while pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD] == pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]:
                                    start = start + 1
                                # combine all dictionaries into one
                                pins_to_combine = pins[temp_start:start + 1]
                                d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                                refactoredDatasheet.append(d)
                                start = start + 1
                                temp_start = start
                            except Exception:
                                # edge case if pin list count is beyond the number of pins
                                pins_to_combine = pins[temp_start:start + 1]
                                d = self.processFunctionProperties(pins_to_combine, macro_state, ignore_blanks)
                                refactoredDatasheet.append(d)
                                start = start + 1
                                temp_start = start
                        # check if the next terminal identifier is different from the present one
                        elif pins[start + 1][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD] != pins[start][datasheetconstants.DATASHEET_TERMINAL_IDENTIFIER_KEYWORD]:
                            refactoredDatasheet.append(pins[start])
                            start = start + 1
                            temp_start = start
                else:
                    start = start + 1
            else:
                refactoredDatasheet.append(pins[start])
                start = start + 1
                temp_start = start
        return refactoredDatasheet

    def processBallMap(self, datasheet, sheetName):
        """

        This method processes the datasheet for ball map files.

        Args:
            datasheet (dict): A dictionary referencing the datasheet from the initial processing
            sheetName (string):  The name of the worksheet being processed

        Returns:

        """
        # Get Pins list
        pins = datasheet[sheetName][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]
        for pin in pins:
            # Apply special processing to each pins list
            if len(pin) != 0:
                interface = pin[datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD][datasheetconstants.DATASHEET_INTERFACE_TYPE]
                all_interfaces = interface.split(" /")
                all_interfaces = list(filter(lambda a: a != ' N/A', all_interfaces))
                pin[datasheetconstants.DATASHEET_NUMBER_OF_SUPPORTED_FUNCTIONS] = len(all_interfaces)
                pin[datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD][datasheetconstants.DATASHEET_INTERFACE_TYPE] = all_interfaces[0]

    def parseSections(self, datasheet, sheetKey, wb, sheetName, sections, map):
        """

        This main processing loop for sections within a worksheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            sheetKey (string):  The dictionary key referencing the output item in the datasheet (output) document
            wb (workbook): The workbook object containing the worksheet.
            sheetName (string): The worksheet name contained in the workbook to parse.
            sections (dict): A dictionary containing the parsing rules for one or more sections within the worksheet.
            map (dict): A map containing the parser rules for the worksheet

        Raises:
            ConfigurationError: An error indicating that the parser rules for the worksheet could not be found.
        """
        try:

            sectionIdx = -1

            if len(sections) > 0:
                self._worksheetSectionIndexWritten = False
                for i in sections:
                    sectionIdx += 1

                    if map.checkIndustryFormat():
                        self.parseSectionWorkGroup(datasheet, sheetKey, wb, sheetName, i, map)
                    else:
                        self.parseSection(datasheet, sheetKey, wb, sheetName, i, map)
                    self.tableCounter += 1
            else:
                s = t('No sections found in map file.  A map file must have at least 1 section')
                d = str(map._mapFileName)
                raise ConfigurationError(s, d)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def applySpecialRules(self, datasheet, map, sheetNames):
        """

        This method applies additional processing to enable output file matches with schema.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            sheetNames (list): List of all sheet names in the file
            map (dict): A map containing the parser rules for the worksheet

        """
        macro_state = map.macroEnabled()
        # Process external components in datasheet
        if datasheetconstants.DATASHEET_PINS_KEYWORD.lower() in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processExternalComponents(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process component id in datasheet
        if datasheetconstants.DATASHEET_COMPONENT_ID in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processComponentID(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process pins and function properties in datasheet
        if datasheetconstants.DATASHEET_PINS_KEYWORD.lower() in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            sheetNamePosition = sheetNames.index(datasheetconstants.DATASHEET_PINS_KEYWORD.lower())
            ignore_blanks = map.ignoreBlanks(sheetNames[sheetNamePosition])
            pins_list = self.processFunctionProperties2(datasheet[spreadsheettypes.SPREADSHEET_TABLES][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()], macro_state, ignore_blanks)
            datasheet[spreadsheettypes.SPREADSHEET_TABLES][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()] = pins_list
            datasheet[spreadsheettypes.SPREADSHEET_TABLES][datasheetconstants.DATASHEET_PINS_KEYWORD.lower()] = self.processConnector(datasheet, spreadsheettypes.SPREADSHEET_TABLES, macro_state, ignore_blanks)
        # Process conditional property in datasheet
        conditional_parameters = self.getCommonOwnerParameters(datasheet, map, datasheetconstants.DATASHEET_CONDITIONAL_PROPERTY)
        if any(parameter in datasheet[spreadsheettypes.SPREADSHEET_TABLES] for parameter in conditional_parameters):
            datasheet = self.processConditionalProperty(datasheet[spreadsheettypes.SPREADSHEET_TABLES], map, conditional_parameters)
        # Process parameters that have type of array of strings in datasheet
        if any(parameter in datasheet[spreadsheettypes.SPREADSHEET_TABLES] for parameter in datasheetconstants.DATASHEET_ARRAY_STRINGS):
            datasheet = self.processArrayStrings(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process graphs in datasheet
        graph_parameters = self.getCommonOwnerParameters(datasheet, map, datasheetconstants.DATASHEET_GRAPH)
        if any(parameter in datasheet[spreadsheettypes.SPREADSHEET_TABLES] for parameter in graph_parameters):
            datasheet = self.processGraph(datasheet[spreadsheettypes.SPREADSHEET_TABLES], graph_parameters)
        # Process integrated fet properties in datasheet
        if any(parameter in datasheet[spreadsheettypes.SPREADSHEET_TABLES] for parameter in datasheetconstants.DATASHEET_POWER_FET_PROPERTIES_PARAMETERS):
            datasheet = self.processIntegratedFetProperties(datasheet[spreadsheettypes.SPREADSHEET_TABLES], datasheetconstants.DATASHEET_POWER_FET_PROPERTIES_PARAMETERS)
        # Process part pin paths in datasheet
        if datasheetconstants.DATASHEET_PART_PIN_PATHS in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processPartPinPaths(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process pin paths in datasheet pins
        if datasheetconstants.DATASHEET_PINS_KEYWORD.lower() in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processPinPaths(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process component protection threshold in datasheet
        if datasheetconstants.DATASHEET_COMPONENT_PROTECTION_THRESHOLD in datasheet[spreadsheettypes.SPREADSHEET_TABLES] and datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processComponentProtectionThresholds(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        # Process power component definitions in datasheet
        if datasheetconstants.DATASHEET_POWER_COMPONENT_DEFINITIONS in datasheet[spreadsheettypes.SPREADSHEET_TABLES]:
            datasheet = self.processPowerComponentDefinitions(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        datasheet = self.processCoreProperties(datasheet[spreadsheettypes.SPREADSHEET_TABLES])
        return datasheet

    def processCoreProperties(self, datasheet):
        """

        This method adds the core properties to the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): updated dictionary with core properties processed

        """
        root_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_ROOT_SCHEMA_NAME)
        root_keys = list(root_schema[schema_constants.PROPERTIES].keys())
        datasheet_keys = list(datasheet.keys())
        output_dict = {}
        modified_datasheet = {}
        modified_datasheet[datasheetconstants.DATASHEET_CORE_PROPERTIES] = {}
        for datasheet_key in datasheet_keys:
            update_value = [(datasheet_key, datasheet[datasheet_key])]
            if datasheet_key not in root_keys:
                modified_datasheet[datasheetconstants.DATASHEET_CORE_PROPERTIES].update(update_value)
            else:
                modified_datasheet.update(update_value)
        # move componentID to top of dictionary
        i, j = 0, 1
        dictionary_tuples = list(modified_datasheet.items())
        for k in range(len(dictionary_tuples)):
            if dictionary_tuples[k][0] == datasheetconstants.DATASHEET_COMPONENT_ID:
                j = k
                break
        dictionary_tuples[i], dictionary_tuples[j] = dictionary_tuples[j], dictionary_tuples[i]
        final_output_dict = dict(dictionary_tuples)
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = final_output_dict
        return output_dict

    def processPowerComponentDefinitions(self, datasheet):
        """

        This method processes the power component definition section of the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): updated dictionary with power component definitions processed

        """
        output_dict = {}
        # Iterate through all power component definition values in dictionary
        for i in range(len(datasheet[datasheetconstants.DATASHEET_POWER_COMPONENT_DEFINITIONS])):
            pcp_element = datasheet[datasheetconstants.DATASHEET_POWER_COMPONENT_DEFINITIONS][i]
            external_component_list = []
            pcp = pcp_element[datasheetconstants.DATASHEET_INSTANCE_DEFINITION]
            if len(pcp) > 0:
                multiple_external_component_index = pcp[0].split(common_constants.COMMA)
                if len(multiple_external_component_index) > 1:
                    multiple_external_component_list = []
                    for component in multiple_external_component_index:
                        external_component_index = component.split(common_constants.DASH)[1]
                        multiple_external_component_list.append(datasheet[datasheetconstants.DATASHEET_INSTANCE_DEF_EXTERNAL_FILE][int(external_component_index) - 4])
                        datasheet[datasheetconstants.DATASHEET_POWER_COMPONENT_DEFINITIONS][i][datasheetconstants.DATASHEET_INSTANCE_DEFINITION] = multiple_external_component_list
                else:
                    external_component_index = pcp[0].split(common_constants.DASH)[1]
                    external_component_list.append(datasheet[datasheetconstants.DATASHEET_INSTANCE_DEF_EXTERNAL_FILE][int(external_component_index) - 4])
                    datasheet[datasheetconstants.DATASHEET_POWER_COMPONENT_DEFINITIONS][i][datasheetconstants.DATASHEET_INSTANCE_DEFINITION] = external_component_list
        del datasheet[datasheetconstants.DATASHEET_INSTANCE_DEF_EXTERNAL_FILE]
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processComponentProtectionThresholds(self, datasheet):
        """

        This method processes the external components section of the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): updated dictionary with component protection thresholds processed

        """
        output_dict = {}
        # Loop through power supply protection in component protection thresholds
        if datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION in datasheet[datasheetconstants.DATASHEET_COMPONENT_PROTECTION_THRESHOLD]:
            for psp in datasheet[datasheetconstants.DATASHEET_COMPONENT_PROTECTION_THRESHOLD][datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION]:
                external_component_list = []
                if len(psp) > 0:
                    multiple_external_component_index = psp.split(common_constants.COMMA)
                    # Process if multiple power supply protection
                    if len(multiple_external_component_index) > 1:
                        multiple_external_component_list = []
                        for component in multiple_external_component_index:
                            external_component_index = component.split(common_constants.DASH)[1]
                            multiple_external_component_list.append(datasheet[datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION][int(external_component_index) - 4])
                            datasheet[datasheetconstants.DATASHEET_COMPONENT_PROTECTION_THRESHOLD][datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION] = multiple_external_component_list
                    else:
                        # Process if single power supply protection
                        external_component_index = psp.split(common_constants.DASH)[1]
                        external_component_list.append(datasheet[datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION][int(external_component_index) - 4])
                        datasheet[datasheetconstants.DATASHEET_COMPONENT_PROTECTION_THRESHOLD][datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION] = external_component_list
        # Delete power supply protection from datasheet
        del datasheet[datasheetconstants.DATASHEET_POWER_SUPPLY_PROTECTION]
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processArrayStrings(self, datasheet):
        """

        This method processes the keys containing array of strings in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): updated dictionary with keys containing array of strings processed

        """
        output_dict = {}
        # Search if any parameter in the datasheetconstants.DATASHEET_ARRAY_STRINGS list is in the datasheet
        for parameter in datasheetconstants.DATASHEET_ARRAY_STRINGS:
            if parameter in datasheet:
                if common_constants.COMMA in datasheet[parameter][0]:
                    component_text = datasheet[parameter][0]
                    # Separate values into list of strings
                    datasheet[parameter] = component_text.split(common_constants.COMMA)
        # Save modified datasheet
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processPartPinPaths(self, datasheet):
        """

        This method processes keys that contain part pins in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): dictionary with Part Pin Paths processed
        """
        output_dict = {}
        # Check for part pin path in datasheet
        if datasheetconstants.DATASHEET_PART_PIN_PATHS in datasheet:
            for datum in datasheet[datasheetconstants.DATASHEET_PART_PIN_PATHS]:
                # Convert string information to list
                if isinstance(datum[datasheetconstants.DATASHEET_COMPONENT_PIN_NAMES][0], str):
                    component_text = datum[datasheetconstants.DATASHEET_COMPONENT_PIN_NAMES][0]
                    datum[datasheetconstants.DATASHEET_COMPONENT_PIN_NAMES] = component_text.split(common_constants.COMMA)
        for pin in datasheet[datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]:
            external_component_list = []
            # Check for part pin path in pin list
            if datasheetconstants.DATASHEET_PIN_PATHS in pin:
                # Get Part Pin Path string
                get_reference = pin[datasheetconstants.DATASHEET_PIN_PATHS][datasheetconstants.DATASHEET_PART_PIN_PATHS]
                component_text = get_reference[0]
                if len(component_text) > 0:
                    # Get list of part pin paths
                    multiple_external_component_index = component_text.split(common_constants.COMMA)
                    if len(multiple_external_component_index) > 1:
                        multiple_external_component_list = []
                        # Process each part pin path
                        for component in multiple_external_component_index:
                            external_component_index = component.split(common_constants.DASH)[1]
                            multiple_external_component_list.append(datasheet[datasheetconstants.DATASHEET_PART_PIN_PATHS][int(external_component_index) - 4])
                            pin[datasheetconstants.DATASHEET_PIN_PATHS][datasheetconstants.DATASHEET_PART_PIN_PATHS] = multiple_external_component_list
                    else:
                        # Process single part pin path
                        external_component_index = component_text.split(common_constants.DASH)[1]
                        external_component_list.append(datasheet[datasheetconstants.DATASHEET_PIN_PATHS][int(external_component_index) - 4])
                        pin[datasheetconstants.DATASHEET_PIN_PATHS] = external_component_list
        # Delete part pin path from dictionary
        del datasheet[datasheetconstants.DATASHEET_PART_PIN_PATHS]
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processIntegratedFetProperties(self, datasheet, power_fet_parameters):
        """

        This method processes keys that contain integrated fet and power fet in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            power_fet_parameters (list): properties linked to power fet parameters

        Returns:
            output_dict (dict): dictionary with integrated fet properties processed
        """
        output_dict = {}
        datasheet[datasheetconstants.DATASHEET_INTEGRATED_FET] = {}
        # Search if any parameter in datasheetconstants.DATASHEET_POWER_FET_PARAMETERS is in datasheet
        for parameter in power_fet_parameters:
            if parameter in datasheet:
                datasheet[datasheetconstants.DATASHEET_INTEGRATED_FET][parameter] = {}
                datasheet[datasheetconstants.DATASHEET_INTEGRATED_FET][parameter].update(datasheet[parameter])
        for datasheet_key in list(datasheet.keys()):
            if datasheet_key in power_fet_parameters:
                del datasheet[datasheet_key]
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processGraph(self, datasheet, graph_parameters):
        """

        This method processes keys that contain graphs in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            graph_parameters (list): properties linked to common/graphProperties

        Returns:
            output_dict (dict): dictionary with component ID processed

        """
        output_dict = {}
        # Search if any parameter in datasheetconstants.DATASHEET_GRAPH_KEY_OWNERS is in datasheet
        for parameter in graph_parameters:
            if parameter in datasheet:
                if datasheetconstants.DATASHEET_CURVE in datasheet[parameter]:
                    for datum in datasheet[parameter][datasheetconstants.DATASHEET_CURVE]:
                        if common_constants.COMMA in datum[datasheetconstants.DATASHEET_X_DATA][0]:
                            # Split value into list from string
                            component_text = datum[datasheetconstants.DATASHEET_X_DATA][0]
                            datum[datasheetconstants.DATASHEET_X_DATA] = [int(i) for i in component_text.split(common_constants.COMMA)]
                        if common_constants.COMMA in datum[datasheetconstants.DATASHEET_Y_DATA][0]:
                            # Split value into list from string
                            component_text = datum[datasheetconstants.DATASHEET_Y_DATA][0]
                            datum[datasheetconstants.DATASHEET_Y_DATA] = [int(i) for i in component_text.split(common_constants.COMMA)]
        # Update dictionary
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def getCommonOwnerParameters(self, datasheet, map, common_owner):
        """

        This method returns all top level keys that link to  property being searched for in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            map (dict): map file json

        Returns:
            output_dict (dict): dictionary with conditional property processed

        """
        output_list = []
        component_type = map.getComponentType()
        schema = readWorkgroupSchema(componentType=component_type)
        schema_keys = list(schema[component_type][schema_constants.PROPERTIES].keys())
        for parameter in schema_keys:
            # Search if any parameter in datasheetconstants.DATASHEET_CONDITIONAL_PROPERTY_KEY_OWNERS is in datasheet
            parameter_type = self.get_reference_value(schema, component_type, parameter)
            if parameter_type == common_owner:
                output_list.append(parameter)
        return output_list

    def processConditionalProperty(self, datasheet, map, conditional_parameters):
        """

        This method processes the conditional properties section of the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet
            map (dict): map file json
            conditional_parameters (list): properties linked to common/conditionalProperties

        Returns:
            output_dict (dict): dictionary with conditional property processed

        """
        output_dict = {}
        component_type = map.getComponentType()
        schema = readWorkgroupSchema(componentType=component_type)
        for parameter in conditional_parameters:
            # Search if any parameter in datasheetconstants.DATASHEET_CONDITIONAL_PROPERTY_KEY_OWNERS is in datasheet
            if parameter in datasheet:
                parameter_type = self.get_reference_value(schema, component_type, parameter)
                if parameter_type == datasheetconstants.DATASHEET_CONDITIONAL_PROPERTY:
                    if datasheetconstants.DATASHEET_CONDITIONS in datasheet[parameter]:
                        if common_constants.COMMA in datasheet[parameter][datasheetconstants.DATASHEET_CONDITIONS][0]:
                            # Split value into list from string
                            component_text = datasheet[parameter][datasheetconstants.DATASHEET_CONDITIONS][0]
                            datasheet[parameter][datasheetconstants.DATASHEET_CONDITIONS] = component_text.split(common_constants.COMMA)
        # Update dictionary
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processComponentID(self, datasheet):
        """

        This method processes the component ID section of the datasheet to enable schema match.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): dictionary with component ID processed
        """
        output_dict = {}
        # Process compliance list in component ID
        if datasheetconstants.DATASHEET_COMPLIANCE_LIST in datasheet[datasheetconstants.DATASHEET_COMPONENT_ID]:
            if common_constants.COMMA in datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_COMPLIANCE_LIST][0]:
                component_text = datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_COMPLIANCE_LIST][0]
                datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_COMPLIANCE_LIST] = component_text.split(common_constants.COMMA)
        # Process orderable MPN in component ID
        if datasheetconstants.DATASHEET_ORDERABLE_MPN in datasheet[datasheetconstants.DATASHEET_COMPONENT_ID]:
            if isinstance(datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN][0], str):
                if common_constants.COMMA in datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN][0]:
                    component_text = datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN][0]
                    datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN] = component_text.split(common_constants.COMMA)
            elif isinstance(datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN][0], int):
                component_text = datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN][0]
                datasheet[datasheetconstants.DATASHEET_COMPONENT_ID][datasheetconstants.DATASHEET_ORDERABLE_MPN] = [str(component_text)]
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processPinPaths(self, datasheet):
        """

        This method processes keys that contain pin paths in the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): dictionary with pin paths processed

        """
        output_dict = {}
        # Iterate through each pin in the datasheet
        for pin in datasheet[datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]:
            # Check if external component key in pin
            if datasheetconstants.DATASHEET_PIN_PATHS in pin:
                # Get the reference to the external component sheet in a list
                pin_part_pin_path_reference = pin[datasheetconstants.DATASHEET_PIN_PATHS]
                for i in range(len(pin_part_pin_path_reference[datasheetconstants.DATASHEET_PART_PIN_PATHS])):
                    pin_part_pin_path_reference[datasheetconstants.DATASHEET_PART_PIN_PATHS] = self.processPartPinPathValues(pin_part_pin_path_reference[datasheetconstants.DATASHEET_PART_PIN_PATHS][i])
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processPartPinPathValues(self, pin_part_pin_path_reference_details):
        part_pin_path_list = []
        part_pin_path_dict = {}
        pin_part_pin_path_reference_details_list = pin_part_pin_path_reference_details.split(common_constants.COMMA)
        for i in range(len(pin_part_pin_path_reference_details_list)):
            start_index = 0
            for j in range(len(pin_part_pin_path_reference_details_list[i])):
                if pin_part_pin_path_reference_details_list[i][j] == common_constants.COLON:
                    part_component_name = pin_part_pin_path_reference_details_list[i][start_index:j].strip()
                    part_component_name = part_component_name[0].lower() + part_component_name[1:]
                    part_component_name = self.format.convert_to_camel_case(part_component_name, True)
                    part_component_value = pin_part_pin_path_reference_details_list[i][j + 1:len(pin_part_pin_path_reference_details_list[i])].strip()
                    if part_component_name == datasheetconstants.DATASHEET_COMPONENT_PIN_NAMES:
                        part_component_value = [x.strip() for x in part_component_value.split(common_constants.SEMI_COLON)]
                        part_pin_path_dict[part_component_name] = part_component_value
                        part_pin_path_list.append(part_pin_path_dict)
                    else:
                        part_pin_path_dict[part_component_name] = part_component_value
                    break
        return part_pin_path_list

    def processExternalComponents(self, datasheet):
        """

        This method processes the external components section of the datasheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet

        Returns:
            output_dict (dict): dictionary with external components processed

        """
        output_dict = {}
        # Iterate through each pin in the datasheet
        for pin in datasheet[datasheetconstants.DATASHEET_PINS_KEYWORD.lower()]:
            # Check if external component key in pin
            if datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS in pin:
                # Get the reference to the external component sheet in a list
                pin_external_component_reference = pin[datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS]
                for i in range(len(pin_external_component_reference)):
                    if datasheetconstants.DATASHEET_CONDITIONS in pin_external_component_reference[i]:
                        pin_external_component_reference[i][datasheetconstants.DATASHEET_CONDITIONS] = self.processConditionsInValues(pin_external_component_reference[i][datasheetconstants.DATASHEET_CONDITIONS])
                if len(pin_external_component_reference) == 1:
                    pin_external_component_reference[i] = self.processValuesOptionsForConditions(pin_external_component_reference[i])
                    pin[datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS] = pin_external_component_reference[i]
                else:
                    pin_external_component_reference = self.processValuesOptionsForConditions(pin_external_component_reference)
                    pin[datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS] = pin_external_component_reference
        output_dict[spreadsheettypes.SPREADSHEET_TABLES] = datasheet
        return output_dict

    def processValuesOptionsForConditions(self, pin_values):
        """

        This method processes values options for conditions in the datasheet.

        Args:
            pin_values (list, dict): A list or dictionary referencing the input pin

        Returns:
            processed_output (list): processed condition values
        """
        value_dictionary = {}
        merge_external_component_dictionaries = {}
        processed_output = []
        if isinstance(pin_values, list):
            for d in pin_values:
                merge_external_component_dictionaries.update(d)
            pin_values = merge_external_component_dictionaries
        for external_component_key in pin_values:
            if external_component_key != datasheetconstants.DATASHEET_COMPONENT_TYPE and external_component_key != datasheetconstants.DATASHEET_CONFIGURATION:
                value_dictionary[external_component_key] = pin_values[external_component_key]
        pin_values[datasheetconstants.DATASHEET_VALUES_PARAMETER] = []
        pin_values[datasheetconstants.DATASHEET_VALUES_PARAMETER].append(value_dictionary)
        for external_component_key in list(pin_values.keys()):
            if external_component_key != datasheetconstants.DATASHEET_VALUES_PARAMETER and external_component_key != datasheetconstants.DATASHEET_COMPONENT_TYPE and external_component_key != datasheetconstants.DATASHEET_CONFIGURATION:
                del pin_values[external_component_key]
        if isinstance(pin_values, dict):
            processed_output.append(pin_values)
            return processed_output
        else:
            return pin_values

    def processConditionsInValues(self, condition_value):
        """

        This method processes conditions in values in the datasheet.

        Args:
            condition_value (list): A list of the conditions separated by a semi colon

        Returns:
            conditions_split (list): list of split strings based on a semi colon
        """
        if len(condition_value) == 1:
            conditions_split = condition_value[0].split(common_constants.SEMI_COLON)
        else:
            conditions_split = condition_value
        return conditions_split

    def processFunctionProperties2(self, pins_list, macro_state, ignore_blanks):
        """

        This method processes the function properties into a list in a hardware connector.

        Args:
            pin_list (list): A list of pins dictionaries with same terminal identifier number which are to be merged to one dictionary
            macro_state (boolean): Flag for if the Excel file is a macro enabled file or not
            ignore_blanks (boolean): Flag for if blanks in the Excel file should be ignored or not

        Returns:
            new_pin_list (list): list with function properties processed.

        """

        new_pin_list = []
        for pin in pins_list:
            d = {}
            for k in pin.keys():
                # store all values asides the function properties in the dictionary
                if k.lower() != datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD.lower():
                    d[k] = pin[k]
                else:
                    # merge all function property values into a list and add to dictionary
                    function_list = [{}]
                    for dicta in pin[k]:
                        dicta_key = list(dicta.keys())[0]
                        dicta_value = dicta[dicta_key]
                        function_list[0][dicta_key] = dicta_value
                    d[datasheetconstants.DATASHEET_FUNCTION_PROPERTIES_KEYWORD] = function_list
            new_pin_list.append(d)
        if macro_state is True:
            d = self.processUnits(d, ignore_blanks)
        return new_pin_list

    def getDatasheetRowIndex(self, datasheet, rowObj, section, map):
        """
        Returns the index value from a row object containing values read from the worksheet.

        Args:
            datasheet (dict): A dictionary referencing the output datasheet.
            rowObj (object): Contains a row of cells/values read from the spreadsheet.
            section(dict): A dictionary containing the parser rules for the current section.
            map (dict): A map containing the parser rules for the worksheet
        """

        try:
            indexOn = map.getIndexOn(section)

            if indexOn is not None:
                # indexOnRow = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                indexOnCol = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]

                # get the column value
                val = self.getColValueFromRow(rowObj, indexOnCol)
                # ExceptionLogger.logDebug(__name__,"Index On Value:  ", val)
                return val

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def _rowHasData(self, rowValues, indexColLetter):
        """
        Returns a boolean indicating whether  the row has any data in it beyond what might be a section header

        Args:
            rowValues:  An array containing row values
            indexColLetter:  The letter of an the indexed column

        """

        hasData = False
        valueCount = 0

        try:

            if rowValues is not None:
                # itemCount = len(rowValues)

                # First see if there is a value in the indexed field.  If not, then the entire row should be skipped
                colIdx = SpreadsheetMap.getColumnIndex(indexColLetter)
                indexColVal = rowValues[colIdx]

                if indexColVal is not None:
                    for i in rowValues:
                        if hasattr(i, "column_letter"):
                            if i.column_letter == indexColLetter:
                                continue
                            else:
                                if i.value is not None:
                                    valueCount += 1

                if valueCount > 0:
                    hasData = True

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

        return hasData

    def serializeIndexWG(self, datasheet, sheetKey, wb, sheetName, section, map):
        """
        Writes the index values to the datasheet output file (in-memory).

        Args:
            datasheet (dict): Dictionary object containing the output datasheet.
            sheetKey (string): A string containing the dictionary key to use for writing index.
            wb (object): Workbook object.
            sheetName (string): Contains the current worksheet name.
            section (dict): Contains the parser rules for the current section for the specified worksheet.
            map (dict): Processing rules for the workbook.
        """

        try:
            indexOn = map.getIndexOn(section)
            indexOnRow = section[spreadsheettypes.SPREADSHEET_MAP_FIELDHEADERS_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
            indexColumnForData = section[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD][0]

            if indexOn is not None:
                dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, indexOn)
                maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, indexOn)
                indexFieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb, sheetName, indexOnRow, indexColumnForData))
                tagList = map.getIncludeTagsDictionaryList(section)

                rowNum = dataStartRow - 1

                indexOnDict = dict()

                for _ in range(rowNum, (rowNum + maxRows) + 1):

                    rowValues = self.getRow(wb, sheetName, rowNum)
                    val = self.getColValueFromRow(rowValues, indexColumnForData, wb, rowNum)

                    rowHasData = self._rowHasData(rowValues, indexColumnForData)

                    # See if this might be a section header in the spreadsheet
                    if (rowHasData) and (val is None):
                        # This is a section header in the spreadsheet
                        rowNum += 1
                        continue
                    elif ((not rowHasData) and ((val is None) or (val == 'defaultValue'))):
                        # End of spreadsheet
                        break
                    else:
                        # Normal row
                        indexOnDict = dict()
                        indexOnDict[indexFieldName] = val

                        if tagList is not None and len(tagList) > 0:

                            for tag in tagList:

                                fieldLabel = self.getCellValue(wb, sheetName,
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                                fieldValue = self.getCellValue(wb, sheetName,
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                                fieldLabel = JsonDataSheet.generateValidJsonFieldName(fieldLabel)

                                if fieldLabel is None:
                                    fieldLabel = "Value not Specified, None"
                                if fieldValue is None:
                                    fieldValue = "Value not Specified, None"
                                if fieldLabel is not None and fieldValue is not None:
                                    indexOnDict[fieldLabel] = fieldValue

                            datasheet[sheetKey].append(indexOnDict)
                            rowNum += 1

                        else:
                            rowNum += 1

                self._worksheetSectionIndexWritten = True

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def serializeIndex(self, datasheet, sheetKey, wb, sheetName, section, map):
        """
        Writes the index values to the datasheet output file (in-memory).

        Args:
            datasheet (dict): Dictionary object containing the output datasheet.
            sheetKey (string): A string containing the dictionary key to use for writing index.
            wb (object): Workbook object.
            sheetName (string): Contains the current worksheet name.
            section (dict): Contains the parser rules for the current section for the specified worksheet.
            map (dict): Processing rules for the workbook.
        """

        try:

            # ExceptionLogger.logDebug(__name__,"!=!=!=SerializeIndex Called, datasheet=",datasheet)

            # get the indexOn field that specifies the data orientation
            indexOn = map.getIndexOn(section)
            indexOnRow = section[spreadsheettypes.SPREADSHEET_MAP_FIELDHEADERS_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
            indexColumnForData = section[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD][0]

            if indexOn is not None:
                dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, indexOn)
                maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, indexOn)
                indexFieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb, sheetName, indexOnRow, indexColumnForData))
                tagList = map.getIncludeTagsDictionaryList(section)

                rowNum = dataStartRow - 1

                indexOnDict = dict()
                # dsRowIdx = -1

                for _ in range(rowNum, (rowNum + maxRows) + 1):

                    rowValues = self.getRow(wb, sheetName, rowNum)
                    val = self.getColValueFromRow(rowValues, indexColumnForData, wb, rowNum)

                    rowHasData = self._rowHasData(rowValues, indexColumnForData)

                    # See if this might be a section header in the spreadsheet
                    if (rowHasData) and (val is None):
                        # This is a section header in the spreadsheet
                        rowNum += 1
                        continue
                    elif ((not rowHasData) and ((val is None) or (val == 'defaultValue'))):
                        # End of spreadsheet
                        break
                    else:
                        # Normal row
                        indexOnDict = dict()
                        indexOnDict[indexFieldName] = val

                        if tagList is not None and len(tagList) > 0:

                            for tag in tagList:

                                fieldLabel = self.getCellValue(wb, sheetName,
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                                fieldValue = self.getCellValue(wb, sheetName,
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                               tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                                fieldLabel = JsonDataSheet.generateValidJsonFieldName(fieldLabel)

                                if fieldLabel is None:
                                    fieldLabel = "Value not Specified, None"
                                if fieldValue is None:
                                    fieldValue = "Value not Specified, None"
                                if fieldLabel is not None and fieldValue is not None:
                                    indexOnDict[fieldLabel] = fieldValue

                            # fieldLabelCheck1 = fieldLabel in indexOnDict
                            # try:
                            #     fieldLabelCheck2 = fieldLabel in datasheet[sheetKey][rowNum - dataStartRow -1]
                            # except:
                            #     pass

                            datasheet[sheetKey].append(indexOnDict)
                            rowNum += 1

                        else:
                            # datasheet[sheetKey].append(indexOnDict)
                            if map.onlyTableName(sheetName):
                                datasheet[sheetName][sheetKey].append({})
                            else:
                                datasheet[sheetKey].append({})
                            rowNum += 1

                self._worksheetSectionIndexWritten = True

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getDatasheetDict(self, datasheet, sheetKey, key, value):
        """
        Returns an object (dictionary or list) that contains the nodes for the datasheet sub-document referenced in the key/value parameters.  This is used
        to look up existing indices.

        Args:
            datasheet (dict): The output datasheet.
            sheetKey (string): Key name for the worksheet subtree to retrieve.
            key (string): Key name for the sub-item to return from the worksheet subtree.
            value (object): The value object for lookup.

        Returns:
            _type_: _description_
        """
        try:

            sheet = datasheet[sheetKey]

            if isinstance(sheet, list):
                for i in sheet:
                    v = i[key]
                    if v == value:
                        return i

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getDatasheetDictGroupBy(self, datasheet, sheetKey, key, value, subHeader):
        try:

            sheet = datasheet[sheetKey]

            if isinstance(sheet, list):
                for i in sheet:
                    v = i[key]
                    if v == value and subHeader not in i:
                        return i

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def get_ref_and_defs(self, json_value, dict_key_name):
        """
        Gets portion of schema that contains references and definitions.

        Args:
            json_value (dict): JSON schema.
            dict_key_name (str): name of high level key that gives access to the properties of the component

        Returns:
            ref_and_def_dictionary (dict): dictionary containing mapping to references and definitions
        """
        ref_and_def_dictionary = {schema_constants.REF: [], schema_constants.DEF_SINGULAR: []}
        # Iterate through keys in schema properties
        if dict_key_name is not None and dict_key_name in json_value:
            try:
                iteration_list = list(json_value[dict_key_name][schema_constants.PROPERTIES].keys())
            except KeyError:
                if dict_key_name == datasheetconstants.DATASHEET_CORE_PROPERTIES:
                    return ref_and_def_dictionary
        else:
            iteration_list = list(json_value[schema_constants.PROPERTIES].keys())
        for key_property in iteration_list:
            # Check if REF in top level key name and its properties
            if dict_key_name is not None and dict_key_name in json_value:
                if schema_constants.REF in json_value[dict_key_name][schema_constants.PROPERTIES][key_property]:
                    if json_value[dict_key_name][schema_constants.PROPERTIES][key_property][schema_constants.REF].startswith(schema_constants.SEARCH_DEF):
                        # dicta[schema_constants.DEF_SINGULAR].append(key_property)
                        ref_and_def_dictionary[schema_constants.DEF_SINGULAR].append(self.get_reference_value(json_value, dict_key_name, key_property))
                    else:
                        ref_and_def_dictionary[schema_constants.REF].append(key_property)
                # Check if REF or DEF in ITEMS in property
                if schema_constants.ITEMS in list(json_value[dict_key_name][schema_constants.PROPERTIES][key_property].keys()):
                    if schema_constants.REF in json_value[dict_key_name][schema_constants.PROPERTIES][key_property][schema_constants.ITEMS]:
                        if json_value[dict_key_name][schema_constants.PROPERTIES][key_property][schema_constants.ITEMS][schema_constants.REF].startswith(schema_constants.SEARCH_DEF):
                            def_value = json_value[dict_key_name][schema_constants.PROPERTIES][key_property][schema_constants.ITEMS][schema_constants.REF].split('/')[-1]
                            ref_and_def_dictionary[schema_constants.DEF_SINGULAR].append(def_value)
                        else:
                            ref_and_def_dictionary[schema_constants.REF].append(key_property)
            else:
                # Check if REF in property only
                if schema_constants.REF in json_value[schema_constants.PROPERTIES][key_property]:
                    if json_value[schema_constants.PROPERTIES][key_property][schema_constants.REF].startswith(schema_constants.SEARCH_DEF):
                        ref_and_def_dictionary[schema_constants.DEF_SINGULAR].append(key_property)
                    else:
                        ref_and_def_dictionary[schema_constants.REF].append(key_property)
                # Check if REF or DEF in ITEMS in property
                if schema_constants.ITEMS in list(json_value[schema_constants.PROPERTIES][key_property].keys()):
                    if schema_constants.REF in json_value[schema_constants.PROPERTIES][key_property][schema_constants.ITEMS]:
                        if json_value[schema_constants.PROPERTIES][key_property][schema_constants.ITEMS][schema_constants.REF].startswith(schema_constants.SEARCH_DEF):
                            ref_and_def_dictionary[schema_constants.DEF_SINGULAR].append(key_property)
                        else:
                            ref_and_def_dictionary[schema_constants.REF].append(key_property)
        return ref_and_def_dictionary

    def name_split(self, ref_name):
        """
        Split link name to obtain file name only.

        Args:
            ref_name (str): name of link to file in schema.

        Returns:
            intermediate_list (str) : name of link file
        """
        if schema_constants.SEARCH_DEF in ref_name:
            intermediate_list = ref_name.split(schema_constants.SEARCH_DEF + "/")
        else:
            intermediate_list = ref_name.split("#/")
        if len(intermediate_list) > 1:
            return intermediate_list[1]
        else:
            intermediate_list = intermediate_list[0].split('/')
            intermediate_list = intermediate_list[-1].split('.')
            return intermediate_list[0]

    def get_reference_value(self, schema, component, parameter):
        """
        Get properties that have a reference.

        Args:
            schema (dict): dictionary of schema containing component
            component (str): name of top level component in the schema.
            parameter (str): name of key to be searched for

        Returns:
            ref_type (str) : reference type
        """
        if schema_constants.REF in schema[component][schema_constants.PROPERTIES][parameter]:
            ref_type = self.name_split(schema[component][schema_constants.PROPERTIES][parameter][schema_constants.REF])
            return ref_type
        else:
            return None

    def get_top_key_name(self, sheetName, schema_json):
        """
        Get properties that have a reference.

        Args:
            schema_json (dict): dictionary of schema containing component
            sheetName (str): name of component in the schema.

        Returns:
            dict_key_name (str) : dictionary top level key for the schema
        """
        if sheetName in datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING:
            dict_key_name = datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[sheetName]
        elif len(list(set(list(schema_json.keys())) - set(datasheetconstants.DATASHEET_DEFAULT_KEYS))) > 0:
            dict_key_name = list(set(list(schema_json.keys())) - set(datasheetconstants.DATASHEET_DEFAULT_KEYS))[0]
        else:
            dict_key_name = sheetName
        return dict_key_name

    def type_search(self, json_value, parameter, full_json={}, root=False, dict_key_name=None):
        """
        Search for property type in JSON schema

        Args:
            json_value (dict): JSON schema.
            parameter (string): parameter to be searched for
            full_json (dict): The dictionary item to be searched.
            root (boolean): If this is the root component or not
            dict_key_name (str) : dictionary top level key for the schema

        Returns:
            response (str) : name of link file
        """
        # Get reference and definitions dictionary
        refs_and_defs = self.get_ref_and_defs(json_value, dict_key_name)
        # Base condition: If no top level dictionary key in JSON schema
        if dict_key_name is None:
            if parameter in list(json_value[schema_constants.PROPERTIES].keys()):
                try:
                    return json_value[schema_constants.PROPERTIES][parameter][schema_constants.TYPE]
                except KeyError:
                    return common_constants.OBJECT_TYPE
        # Base condition: If top level dictionary key in JSON schema
        elif dict_key_name in json_value:
            if schema_constants.ONE_OF in json_value[dict_key_name]:
                return None
            # Base condition: If parameter in immediate list return it or object if KeyError
            if parameter in list(json_value[dict_key_name][schema_constants.PROPERTIES].keys()):
                try:
                    return json_value[dict_key_name][schema_constants.PROPERTIES][parameter][schema_constants.TYPE]
                except KeyError:
                    return common_constants.OBJECT_TYPE
            # Base condition: If parameter in immediate list return it or pass if KeyError
            elif schema_constants.PROPERTIES in json_value:
                try:
                    return json[schema_constants.PROPERTIES][parameter][schema_constants.TYPE]
                except KeyError:
                    pass
            # Base condition: If parameter in immediate list return it or object if KeyError
            elif parameter in list(json_value[dict_key_name].keys()):
                try:
                    return json_value[dict_key_name][parameter][schema_constants.TYPE]
                except KeyError:
                    return common_constants.OBJECT_TYPE
        # Base condition: If parameter is in root schema
        if dict_key_name == datasheetconstants.DATASHEET_ROOT_SCHEMA_NAME and parameter in datasheetconstants.DATASHEET_EXTERNAL_FILE_MAP_KEYS:
            return common_constants.OBJECT_TYPE
        # Base condition: Search for type in properties if top level dictionary not in schema or pass if KeyError
        if dict_key_name not in json_value and schema_constants.PROPERTIES in json_value:
            try:
                return json_value[schema_constants.PROPERTIES][parameter][schema_constants.TYPE]
            except KeyError:
                pass
        # Base condition: Special case where parameter is a hash symbol
        if parameter == '#' or parameter.isdigit():
            return None
        # Base condition: Else recursively search ref and def for parameter
        for refs in refs_and_defs[schema_constants.REF]:
            if parameter == refs:
                return common_constants.OBJECT_TYPE
            if root is True:
                name_key = refs
                if name_key in datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING:
                    name_key = datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[name_key]
                    root = False
            else:
                if dict_key_name in json_value:
                    try:
                        name_key = self.name_split(json_value[dict_key_name][schema_constants.PROPERTIES][refs][schema_constants.REF])
                    except KeyError:
                        name_key = self.name_split(json_value[dict_key_name][schema_constants.PROPERTIES][refs][schema_constants.ITEMS][schema_constants.REF])
                elif schema_constants.PROPERTIES in json_value:
                    try:
                        name_key = self.name_split(json_value[schema_constants.PROPERTIES][refs][schema_constants.REF])
                    except KeyError:
                        if refs in datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING:
                            name_key = datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[refs]
            ideal_json = readWorkgroupSchema(name_key)
            # response = self.type_search(ideal_json[name_key], parameter, full_json=ideal_json, dict_key_name=None)
            response = self.type_search(ideal_json, parameter, full_json=ideal_json, dict_key_name=name_key, root=root)
            if response is not None:
                return response
        for defs in refs_and_defs[schema_constants.DEF_SINGULAR]:
            if bool(full_json) is not False:
                ideal_json = full_json[schema_constants.DEF_PLURAL][defs]
                if datasheetconstants.DATASHEET_NAME_KEY in locals():
                    response = self.type_search(ideal_json, parameter, full_json={}, dict_key_name=name_key, root=root)
                else:
                    response = self.type_search(ideal_json, parameter, full_json={}, dict_key_name=dict_key_name, root=root)
                if response is not None:
                    return response
            elif defs in json_value[schema_constants.DEF_PLURAL]:
                ideal_json = json_value[schema_constants.DEF_PLURAL][defs]
                response = self.type_search(ideal_json, parameter, full_json={}, dict_key_name=None)
                if response is not None:
                    return response

    def processObjectWG(self, datasheet, schema_json, sheetName, map, key, updatedDict, property_dict, entry_index, root=False, valuePresence=False, sub_entry_index=0):
        """
        Process Object type in the schema

        Args:
            datasheet (dict): the output datasheet.
            schema_json (dict): the JSON schema from the workgroup
            sheetName (string): name of sheet being processed
            map (SpreadsheetMap): map class to enable processing of inputs from Map file
            key (string): key to be processed.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
            property_dict (dict): The dictionary item that shows the mapping of names from Excel file to JSON schema.
            entry_index (integer): the row currently being processed in the Excel sheet.

        Returns:
            datasheet (dict) : updated datasheet
        """
        # If sheetName is not equal to component
        if sheetName != map.getComponentType():
            if sheetName not in datasheet:
                datasheet[sheetName] = {}
        dict_key_name = self.getTopLevelDictionaryKeyName(schema_json, sheetName)
        data_type = self.type_search(schema_json, key, dict_key_name=dict_key_name, root=root)
        if (list(updatedDict.keys())[0] in datasheetconstants.DATASHEET_VALUE_KEYS) or (data_type is None and key == 'T'):
            datasheet = self.processArrayWG(datasheet[sheetName], readWorkgroupSchema('values'), 'values', map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, sub_entry_index=sub_entry_index)
        # If data type of key is an object recursively process the object
        elif data_type == common_constants.OBJECT_TYPE:
            if sheetName == map.getComponentType():
                updatedDict = self.processObjectWG(datasheet, schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, root, sub_entry_index=sub_entry_index)
            else:
                updatedDict = self.processObjectWG(datasheet[sheetName], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, root, sub_entry_index=sub_entry_index)
        # If data type of key is an array recursively process the array
        elif data_type == common_constants.ARRAY_TYPE and any(isinstance(i, dict) for i in updatedDict.values()) is False:
            if sheetName not in datasheet:
                updatedDict = self.processArrayWG(datasheet, schema_json, key, map, key, updatedDict[key], property_dict, entry_index, root)
            else:
                updatedDict = self.processArrayWG(datasheet[sheetName], schema_json, key, map, key, updatedDict[key], property_dict, entry_index, root)
        elif data_type == common_constants.ARRAY_TYPE and any(isinstance(i, dict) for i in updatedDict.values()) is True:
            updatedDict = self.processArrayWG(datasheet[sheetName], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, root)
        elif data_type != common_constants.ARRAY_TYPE:
            # property_type = self.type_search(schema_json, key, dict_key_name=map.getComponentType())
            datasheet = self.updateDatasheetWithNewKeyObject(map, datasheet, sheetName, key, data_type, updatedDict, property_dict)
        return datasheet

    def processArrayWG(self, datasheet, schema_json, sheetName, map, key, updatedDict, property_dict, entry_index, root=False, sub_entry_index=0):
        """
        Process Array type in the schema

        Args:
            datasheet (dict): the output datasheet.
            schema_json (dict): the JSON schema from the workgroup
            sheetName (string): name of sheet being processed
            map (SpreadsheetMap): map class to enable processing of inputs from Map file
            key (string): key to be processed.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
            property_dict (dict): The dictionary item that shows the mapping of names from Excel file to JSON schema.
            entry_index (integer): the row currently being processed in the Excel sheet.

        Returns:
            datasheet (dict) : updated datasheet

        """
        # Base condition: If sheetName is not in datasheet
        if sheetName not in datasheet:
            datasheet[sheetName] = []
        dict_key_name = self.getTopLevelDictionaryKeyName(schema_json, sheetName)
        # dict_key_name = list(set(list(schema_json.keys())) - set(datasheetconstants.DATASHEET_DEFAULT_KEYS))[0]
        data_type = self.type_search(schema_json, key, dict_key_name=dict_key_name)
        # Base condition: If updated dict is a string, integer of float
        if isinstance(updatedDict, str) or isinstance(updatedDict, int) or isinstance(updatedDict, float):
            if sheetName == datasheetconstants.DATASHEET_CONDITIONS and isinstance(updatedDict, str):
                condition_value_split = [x.strip() for x in updatedDict.split(common_constants.SEMI_COLON)]
                datasheet[sheetName].extend(condition_value_split)
            else:
                datasheet[sheetName].append(updatedDict)
        # If data type of key is an object recursively process the object
        elif data_type == common_constants.OBJECT_TYPE:
            if isinstance(datasheet[sheetName], list):
                if len(datasheet[sheetName]) == 0 and entry_index == 0:
                    updatedDict = self.processObjectWG({}, schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
                    datasheet[sheetName].append(updatedDict)
                elif 0 <= entry_index >= len(datasheet[sheetName]):
                    updatedDict = self.processObjectWG({}, schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
                    datasheet[sheetName].append(updatedDict)
                elif len(list(updatedDict[key])) > 0:
                    updatedDict = self.processObjectWG(datasheet[sheetName][entry_index], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
                else:
                    updatedDict = self.processObjectWG(datasheet[sheetName][entry_index], schema_json, key, map, list(updatedDict[key]), updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
            else:
                updatedDict = self.processObjectWG(datasheet[sheetName], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index)
        # If data type of key is an array recursively process the array
        elif data_type == common_constants.ARRAY_TYPE:
            if len(list(updatedDict[key])) > 0:
                if dict_key_name == datasheetconstants.DATASHEET_VALUE_OPTIONS_PARAMETER or list(updatedDict.keys())[0] == datasheetconstants.DATASHEET_CONDITIONS:
                    updatedDict = self.processArrayWG(datasheet[sheetName][sub_entry_index], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
                else:
                    updatedDict = self.processArrayWG(datasheet[sheetName][entry_index], schema_json, key, map, list(updatedDict[key])[0], updatedDict[key], property_dict, entry_index, sub_entry_index=sub_entry_index)
            else:
                updatedDict = self.processArrayWG(datasheet[sheetName][entry_index], schema_json, key, map, list(updatedDict[key]), updatedDict[key], property_dict, entry_index)
        else:
            datasheet = self.updateDatasheetWithNewKeyArray(datasheet, sheetName, schema_json, key, dict_key_name, updatedDict, property_dict, entry_index, sub_entry_index)
        return datasheet

    def getTopLevelDictionaryKeyName(self, schema_json, sheetName):
        """
        Finds top level dictionary key name

        Args:
            schema_json (dict): the JSON schema from the workgroup
            sheetName (string): name of sheet being processed


        Returns:
            dict_key_name (string) : top level dictionary key name in schema

        """
        if sheetName in datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING:
            dict_key_name = datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[sheetName]
        elif len(list(set(list(schema_json.keys())) - set(datasheetconstants.DATASHEET_DEFAULT_KEYS))) > 0:
            dict_key_name = list(set(list(schema_json.keys())) - set(datasheetconstants.DATASHEET_DEFAULT_KEYS))[0]
        else:
            dict_key_name = sheetName
        return dict_key_name

    def updateDatasheetWithNewKeyArray(self, datasheet, sheetName, schema_json, key, dict_key_name, updatedDict, property_dict, entry_index, sub_entry_index):
        """
        Update datasheet with new key and value after processing for arrays

        Args:
            datasheet (dict): the output datasheet.
            schema_json (dict): the JSON schema from the workgroup
            sheetName (string): name of sheet being processed
            map (SpreadsheetMap): map class to enable processing of inputs from Map file
            key (string): key to be processed.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
            property_dict (dict): The dictionary item that shows the mapping of names from Excel file to JSON schema.
            entry_index (integer): the row currently being processed in the Excel sheet.
            sub_entry_index (integer): the sub row currently being processed in the Excel sheet.

        Returns:
            datasheet (dict) : updated datasheet

        """
        property_type = self.type_search(schema_json, key, dict_key_name=dict_key_name)
        current_type = type(updatedDict[key]).__name__
        if property_dict[property_type] == current_type:
            pass
        else:
            updatedDict = self.typeConversion(updatedDict, property_dict[property_type], current_type)
        datasheetDict = datasheet
        if (key in datasheet):
            if (datasheetDict[key] != updatedDict[key]):
                prevDict = datasheetDict[key]
                prevDict.update(updatedDict[key])
                updatedDict = prevDict
            datasheet[key] = updatedDict
        else:
            if sub_entry_index > 0 or dict_key_name == datasheetconstants.DATASHEET_VALUE_OPTIONS_PARAMETER:
                if 0 <= sub_entry_index < len(datasheet[sheetName]):
                    datasheet[sheetName][sub_entry_index].update(updatedDict)
                else:
                    datasheet[sheetName].append(updatedDict)
            else:
                if 0 <= entry_index < len(datasheet[sheetName]):
                    datasheet[sheetName][entry_index].update(updatedDict)
                else:
                    datasheet[sheetName].append(updatedDict)
        return datasheet

    def updateDatasheetWithNewKeyObject(self, map, datasheet, sheetName, key, data_type, updatedDict, property_dict):
        """
        Update datasheet with new key and value after processing for objects

        Args:
            datasheet (dict): the output datasheet.
            sheetName (string): name of sheet being processed
            map (SpreadsheetMap): map class to enable processing of inputs from Map file
            key (string): key to be processed.
            data_type (string): data type of the input key to be processed.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
            property_dict (dict): The dictionary item that shows the mapping of names from Excel file to JSON schema.

        Returns:
            datasheet (dict) : updated datasheet

        """
        current_type = type(updatedDict[key]).__name__
        if property_dict[data_type] == current_type:
            pass
        else:
            updatedDict = self.typeConversion(updatedDict, property_dict[data_type], current_type)
        datasheetDict = datasheet
        if (key in datasheet) and (key != datasheetconstants.DATASHEET_PART_TYPE):
            if (datasheetDict[key] != updatedDict[key]):
                prevDict = datasheetDict[key]
                prevDict.update(updatedDict[key])
                updatedDict = prevDict
            datasheet[key] = updatedDict
        else:
            if sheetName != map.getComponentType():
                datasheet[sheetName].update(updatedDict)
            else:
                datasheet.update(updatedDict)
        return datasheet

    def updateDatasheetDictWG(self, datasheet: dict, sheetKey, key: str, value, updatedDict: dict, rowNum, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName, sub_entry_index=0) -> None:
        """
        Updates the datasheet output sub-document with new/updated values. Ensures match with Industry Workgroup Schema

        Args:
            datasheet (dict): the output datasheet.
            sheetKey (string): Key name for the worksheet subtree to update.
            key (string): Key name for the sub-item to return from the worksheet subtree.
            value (dict): The value object for lookup.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
        """
        try:
            componentType = map.getComponentType()
            root_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_ROOT_SCHEMA_NAME)
            defined_schema = readWorkgroupSchema(componentType)
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            # Process presence of a dash in sheetName
            if "-" in sheetName:
                full_name = sheetName.split(common_constants.DASH)
                sheetName = full_name[1]
                sheetName = self.format.convert_to_camel_case(sheetName, industryFormatCheck=True)
            if sheetName == componentType:
                datasheet = self.processObjectWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, sub_entry_index=sub_entry_index)
            else:
                if sheetKey in datasheetconstants.DATASHEET_COMPONENT_KEYS:
                    if self.type_search(root_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName]) == common_constants.OBJECT_TYPE:
                        defined_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetKey])
                        datasheet = self.processObjectWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, root=True)
                    elif self.type_search(root_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName]) == common_constants.ARRAY_TYPE:
                        defined_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetKey])
                        datasheet = self.processArrayWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, root=True, sub_entry_index=sub_entry_index)
                elif sheetKey in datasheetconstants.DATASHEET_EXTERNAL_PINS:
                    if self.type_search(root_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName],
                                        dict_key_name=datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[datasheetconstants.DATASHEET_PINS_KEYWORD]) == common_constants.OBJECT_TYPE:
                        defined_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_SCHEMA_MAPPING[datasheetconstants.DATASHEET_PINS_KEYWORD])
                        datasheet = self.processObjectWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, root=True)
                    elif self.type_search(root_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName],
                                          dict_key_name=datasheetconstants.DATASHEET_COMPONENT_COMMON_MAPPING[datasheetconstants.DATASHEET_PINS_KEYWORD]) == common_constants.ARRAY_TYPE:
                        defined_schema = readWorkgroupSchema(datasheetconstants.DATASHEET_SCHEMA_MAPPING[datasheetconstants.DATASHEET_PINS_KEYWORD])
                        datasheet = self.processArrayWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index)
                else:
                    if self.type_search(defined_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName], dict_key_name=componentType) == common_constants.OBJECT_TYPE:
                        datasheet = self.processObjectWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, sub_entry_index=sub_entry_index)
                    elif self.type_search(defined_schema, datasheetconstants.DATASHEET_SCHEMA_MAPPING[sheetName], dict_key_name=componentType) == common_constants.ARRAY_TYPE:
                        datasheet = self.processArrayWG(datasheet, defined_schema, sheetName, map, key, updatedDict, datasheetconstants.DATASHEET_SCHEMA_PROPERTY_MAPPING, entry_index, sub_entry_index=sub_entry_index)
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def typeConversion(self, updatedDict, property_type, current_type, ignoreBlanks=False):
        """
        Converts type of parameters to ensure match with Workgroup Schema

        Args:
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
            property_type (string): the type of the property in the schema.
            current_type (string): the type of the property after being read from the Excel file.
            ignoreBlanks (boolean): if blanks are ignored in the worksheet or not.

        Returns:
            updatedDict (dict) : updated datasheet with type conversion
        """
        # Convert from string to list
        if property_type == common_constants.LIST_TYPE and current_type == common_constants.STRING_SHORT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            list_value = updatedDict[key].split(common_constants.COMMA)
            updatedDict[key] = list_value
        # Convert from string to integer
        elif property_type == common_constants.INT_SHORT_TYPE and current_type == common_constants.STRING_SHORT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            str_value = str(updatedDict[key])
            if ignoreBlanks is True:
                error_message = "{insert_key} should be a number. Please check".format(insert_key=key)
                ExceptionLogger.logError(__name__, error_message)
            updatedDict[key] = str_value
        # Convert from integer to string
        elif property_type == common_constants.STRING_SHORT_TYPE and current_type == common_constants.INT_SHORT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            str_value = str(updatedDict[key])
            updatedDict[key] = str_value
        # Convert from float to string
        elif property_type == common_constants.STRING_SHORT_TYPE and current_type == common_constants.FLOAT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            str_value = str(updatedDict[key])
            updatedDict[key] = str_value
        # Convert from dictionary to list
        elif property_type == common_constants.LIST_TYPE and current_type == common_constants.DICT_SHORT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            new_dict = {key: [updatedDict[key]]}
            updatedDict = new_dict
        # Convert from string to boolean
        elif property_type == common_constants.BOOLEAN_SHORT_TYPE and current_type == common_constants.STRING_SHORT_TYPE:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            new_dict = {key: False}
            updatedDict = new_dict
        return updatedDict

    def updateDatasheetDict(self, datasheet, sheetKey, key, value, updatedDict, rowNum, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName):
        """
        Updates the datasheet output sub-document with new/updated values.

        Args:
            datasheet (dict): the output datasheet.
            sheetKey (string): Key name for the worksheet subtree to update.
            key (string): Key name for the sub-item to return from the worksheet subtree.
            value (object): The value object for lookup.
            updatedDict (dict): The dictionary item to replace/update in the datasheet.
        """
        try:
            dictKeys = list(updatedDict)
            key = dictKeys[0]
            if map.onlyTableName(sheetName):
                datasheetDict = datasheet[sheetName][sheetKey][entry_index]
                if (key in datasheet[sheetName][sheetKey][entry_index]):
                    if (datasheetDict[key] != updatedDict[key]):
                        prevDict = datasheetDict[key]
                        prevDict.update(updatedDict[key])
                        updatedDict = prevDict
                    datasheet[sheetName][sheetKey][entry_index][key] = updatedDict
                else:
                    datasheet[sheetName][sheetKey][entry_index].update(updatedDict)
            else:
                datasheetDict = datasheet[sheetKey][entry_index]
                if (key in datasheet[sheetKey][entry_index]):
                    if (datasheetDict[key] != updatedDict[key]):
                        prevDict = datasheetDict[key]
                        prevDict.update(updatedDict[key])
                        updatedDict = prevDict
                    datasheet[sheetKey][entry_index][key] = updatedDict
                else:
                    datasheet[sheetKey][entry_index].update(updatedDict)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    # returns a list containing  the field headers for a section and returns a list
    def getFieldHeadersForSection(self, wb, sheetName, section, map):
        """Returns a list containing the field headers for a section.

        Args:
            wb (object): The workbook object.
            sheetName (string): Worksheet name.
            section (dict): The parser rules for the current section.
            map (object): The parser rules for the worksheet.

        Raises:
            ConfigurationError: Error if parser rules are not found.

        Returns:
            list: List containing the field headers for a section.
        """
        try:
            # ExceptionLogger.logDebug(__name__,"section:",section)

            spans = map.getSpansColumns(section)

            # check whether this is item is in a GroupBy.
            groupBy = map.getGroups(section)

            # ExceptionLogger.logDebug(__name__,"spans:",spans)
            # ExceptionLogger.logDebug(__name__,"section:",section)

            if spans is not None:

                spanHeader = []
                fieldHeaderObj = map.getFieldHeaders(section)
                fieldHeaderRow = map.getRow(fieldHeaderObj)

                for i in range(len(spans)):
                    col = spans[i]
                    isMerge = self.isMerged(wb, fieldHeaderRow, col)
                    isSubHeader = False
                    if groupBy is not None:
                        isSubHeader = SpreadsheetMap.columnInGroupByList(col, groupBy)
                    if (isMerge):
                        data = self.getCellValueIfMerged(wb, fieldHeaderRow, col)
                    else:
                        data = self.getCellValue(wb, sheetName, fieldHeaderRow, col)
                    if type(data) is not str:
                        # Normalize that all data to be a string
                        data = str(data)
                    dataStripped = data.strip()
                    if data == serializationconstants.EMPTY_HEADER or dataStripped == "":
                        data = f"empty header col {col}"
                    industryFormatCheck = map.checkIndustryFormat()
                    fieldHeader = self.format.format_name_spreadsheet(data, industryFormatCheck)
                    if isSubHeader:
                        # Build the header list for the sub-document field names to be added
                        groupByItemHeaderDictList = SpreadsheetMap.getGroupByItemHeaderDictList(groupBy)
                        groupByItemHeaders = self.getFieldHeaderList(wb, sheetName, groupByItemHeaderDictList, map)  # get a list of the item headers
                        deepKey = ""
                        groupIndex = 0
                        for g in groupByItemHeaders:
                            groupToValidate = groupBy[groupIndex]
                            groupSpans = groupToValidate[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD]
                            # Only add the grouping for the spans section
                            if col in groupSpans:
                                deepKey = deepKey + g + '-'
                            groupIndex += 1
                        fieldHeader = deepKey + fieldHeader
                    if fieldHeader not in spanHeader:
                        spanHeader.append(fieldHeader)
                    else:
                        spanHeader.append(fieldHeader + '_repeated_' + col + '_' + str(i))

                return spanHeader, fieldHeaderRow
            else:
                # expecting at east one span
                s = t('No spans found in map file.  Map file and a section must have at least 1 span')
                d = map._mapFileName
                raise ConfigurationError(s, d)

            # ExceptionLogger.logDebug(__name__,"spanHeader:",spanHeader)
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getFieldHeaderList(self, wb, sheetName, itemHeaderDictList, map):  # itemHeaderDictList is list of {row:n,col:a} items
        """Returns a list containing field headers for a worksheet.  These headers are used to generate the field names in the resulting datasheet.

        Args:
            wb (object): Workbook object.
            sheetName (string): Current worksheet name.
            itemHeaderDictList (dict): The dictionary containing the field headers.

        Returns:
            list: Contains the field headers.
        """

        json_field_names = []
        try:

            # ExceptionLogger.logDebug(__name__,"itemHeaderDictList=",itemHeaderDictList)
            for i in itemHeaderDictList:
                # fieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb,sheetName,
                #                                                                        i[spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                #                                                                        i[spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]))

                fieldName = self.getCellValue(wb, sheetName, i[spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                              i[spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])

                if spreadsheettypes.SPREADSHEET_MAP_VALUE_COLUMN_FIELD in i:
                    if len(i[spreadsheettypes.SPREADSHEET_MAP_VALUE_COLUMN_FIELD]):
                        fieldValue = self.getCellValue(wb, sheetName, i[spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                       i[spreadsheettypes.SPREADSHEET_MAP_VALUE_COLUMN_FIELD]).strip()
                        # if there is a field value, append
                        if len(fieldValue) > 0:
                            fieldName += fieldValue

                # ExceptionLogger.logDebug(__name__,"fieldName=",fieldName)
                # jsonFieldName = JsonDataSheet.generateValidJsonFieldName(fieldName)
                industryFormatCheck = map.checkIndustryFormat()
                jsonFieldName = self.format.format_name_spreadsheet(fieldName, industryFormatCheck)
                json_field_names.append(jsonFieldName)

            return json_field_names

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def processSpanValuesForIndexedSectionWG(self, datasheet, sheetKey, wb, sheetName, section, spans, fieldHeaders, indexOn, rowNumber, rowValues, map, dataStartOffset, rowNumTracker, entry_index):
        """
        Processing loop for the spans within an indexed section.

        Args:
            datasheet (dict): datasheet output document.
            sheetKey (string): worksheet key.
            wb (object): workbook object_
            sheetName (string): name of current worksheet
            section (dict): describes processing rules within the map file for a section.
            spans (list): list of columns to process in the current section
            fieldHeaders (list): field headers to use in writing datasheet
            indexFieldName (string): field name of the index.
            rowValues (object): a row of values from the worksheet
            map (dict): parser rules for the worksheet
        """
        try:
            if indexOn is not None:
                columnIndex = -1

                # Get processing conditions
                colsToIgnore = map.getColsToIgnore(sheetName)
                ignoreBlanksFlag = map.ignoreBlanks(sheetName)

                # Get processing parameters
                indexOnRow = section[spreadsheettypes.SPREADSHEET_MAP_FIELDHEADERS_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                indexOnCol = section[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD][0]
                headerCol = indexOnCol
                indexFieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb, sheetName, indexOnRow, indexOnCol))

                # Perform Processing
                if spans is not None and len(spans) > 0:
                    for s in spans:
                        if s in colsToIgnore:
                            columnIndex += 1
                            pass
                        else:
                            val = None
                            itemDict = None
                            groupBy = None
                            itemSubNewDict = None
                            columnIndex += 1

                            val = self.getColValueFromRow(rowValues, s, wb, rowNumber)
                            if ignoreBlanksFlag is True and len(str(val).strip()) == 0:
                                pass
                            elif ((val is not None) and (len(str(val)) >= 0)):
                                indexValue = self.getColValueFromRow(rowValues, headerCol, wb, rowNumber)
                                groupBy = map.getGroups(section)
                                if groupBy is not None:

                                    spanListInGroupBy = SpreadsheetMap.getListSpanInGroupBy(section)
                                    if s in spanListInGroupBy and SpreadsheetMap.getSubObjectFlag(groupBy, s):
                                        compoundNameList = fieldHeaders[columnIndex].split('-')
                                        baseHeader = compoundNameList[0]
                                        subHeader = compoundNameList[1]
                                        itemSubNewDict = dict({subHeader: val})
                                        itemDict = dict({baseHeader: itemSubNewDict})
                                    else:
                                        if val is not None and len(str(val)) >= 0:
                                            itemDict = dict({fieldHeaders[columnIndex]: val})
                                else:

                                    if val is not None and len(str(val)) >= 0:
                                        itemDict = dict({fieldHeaders[columnIndex]: val})

                                if itemDict is not None:
                                    # Process Macro
                                    if map.macroEnabled() is True:
                                        itemDict = self.processUnitMacro(map, sheetName, itemDict)
                                        if isinstance(itemDict, dict) and list(itemDict.keys())[0] == datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS:
                                            itemDict = self.processPinExternalComponentsMacro(map, sheetName, itemDict)
                                        if isinstance(itemDict, dict) and isinstance(itemDict[list(itemDict.keys())[0]], dict):
                                            sub_header_1 = list(itemDict.keys())[0]
                                            sub_header_2 = list(itemDict[sub_header_1].keys())[0]
                                            if isinstance(itemDict[sub_header_1][sub_header_2], str) and itemDict[sub_header_1][sub_header_2].startswith(datasheetconstants.DATASHEET_TYP_VALUE):
                                                sheetKey, sheetName = sub_header_1, sub_header_1
                                                itemDict = self.processUnitMacro(map, sheetName, itemSubNewDict)
                                        sub_entry_index = 0
                                        # Process based on input types
                                        if isinstance(itemDict, list) and len(itemDict) != 0:
                                            for k in range(len(itemDict)):
                                                item = itemDict[k]
                                                top_level_key = list(item.keys())[0]
                                                sub_unit_dict = item[top_level_key]
                                                if list(sub_unit_dict.keys())[0] == datasheetconstants.DATASHEET_COMPONENT_TYPE and k > 0 and top_level_key == datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS:
                                                    sub_entry_index = sub_entry_index + 1
                                                elif k > 0 and top_level_key != datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS:
                                                    current_unit_value = list(sub_unit_dict.keys())[0]
                                                    prev_item = itemDict[k - 1]
                                                    prev_top_level_key = list(prev_item.keys())[0]
                                                    prev_sub_unit_dict = prev_item[prev_top_level_key]
                                                    previous_unit_value = list(prev_sub_unit_dict.keys())[0]
                                                    if datasheetconstants.DATASHEET_VALUE_KEYS.index(current_unit_value) < datasheetconstants.DATASHEET_VALUE_KEYS.index(previous_unit_value):
                                                        sub_entry_index = sub_entry_index + 1
                                                self.updateDatasheetDictWG(datasheet, sheetKey, indexFieldName, indexValue, item, rowNumber, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName, sub_entry_index)
                                        elif isinstance(itemDict, dict) and len(itemDict) != 0:
                                            self.updateDatasheetDictWG(datasheet, sheetKey, indexFieldName, indexValue, itemDict, rowNumber, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName)
                                    else:
                                        self.updateDatasheetDictWG(datasheet, sheetKey, indexFieldName, indexValue, itemDict, rowNumber, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def processUnitMacro(self, map, sheetName, itemDict):
        """
        Process unit values specified in the macro file. If non unit values return original item dictionary

        Args:
            map (dict): parser rules for the worksheet
            sheetName (string): name of current worksheet
            itemDict (dict): dictionary containing values for that cell
        """
        all_keys = list(itemDict.keys())
        ignore_blanks = map.ignoreBlanks(sheetName)
        allItemDict = []
        if len(all_keys) > 0:
            dict_value = itemDict[all_keys[0]]
            # Convert unit values from string to dictionary
            if isinstance(dict_value, str):
                if dict_value.startswith(datasheetconstants.DATASHEET_TYP_VALUE):
                    loop_values = self.convertUnitStrToDict(dict_value, ignore_blanks)
                elif all_keys[0] in datasheetconstants.DATASHEET_UNIT_LIST and len(dict_value) == 0 and ignore_blanks is False:
                    dict_value = datasheetconstants.DATASHEET_EMPTY_MACRO_UNIT
                    loop_values = self.convertUnitStrToDict(dict_value, ignore_blanks)
                else:
                    return itemDict
            else:
                return itemDict
        for i in range(len(loop_values)):
            if len(loop_values[i]) != 0:
                # Add unit dictionary to original parameter containing the unit
                for key in loop_values[i]:
                    inner_dict = {}
                    outer_dict = {}
                    inner_dict[key] = loop_values[i][key]
                    outer_dict[all_keys[0]] = inner_dict
                    allItemDict.append(outer_dict)
        return allItemDict

    def processPinExternalComponentsMacro(self, map, sheetName, itemDict):
        """
        Process unit values specified in the macro file. If non unit values return original item dictionary

        Args:
            map (dict): parser rules for the worksheet
            sheetName (string): name of current worksheet
            itemDict (dict): dictionary containing values for that cell
        """
        all_keys = list(itemDict.keys())
        ignore_blanks = map.ignoreBlanks(sheetName)
        allItemDict = []
        if len(all_keys) > 0:
            dict_value = itemDict[all_keys[0]]
            # Convert unit values from string to dictionary
            if isinstance(dict_value, str):
                if dict_value.startswith(datasheetconstants.DATASHEET_PIN_EXTERNAL_COMPONENTS_START):
                    loop_values = self.convertUnitStrToDict(dict_value, ignore_blanks, processorType=datasheetconstants.DATASHEET_EXTERNAL_COMPONENTS)
                elif all_keys[0] in datasheetconstants.DATASHEET_UNIT_LIST and len(dict_value) == 0 and ignore_blanks is False:
                    dict_value = datasheetconstants.DATASHEET_EMPTY_MACRO_UNIT
                    loop_values = self.convertUnitStrToDict(dict_value, ignore_blanks)
                else:
                    return itemDict
            else:
                return itemDict
        for i in range(len(loop_values)):
            if len(loop_values[i]) != 0:
                # Add unit dictionary to original parameter containing the unit
                for key in loop_values[i]:
                    inner_dict = {}
                    outer_dict = {}
                    inner_dict[key] = loop_values[i][key]
                    outer_dict[all_keys[0]] = inner_dict
                    allItemDict.append(outer_dict)
        return allItemDict

    def processSpanValuesForIndexedSection(self, datasheet, sheetKey, wb, sheetName, section, spans, fieldHeaders, indexOn, rowNumber, rowValues, map, dataStartOffset, rowNumTracker, entry_index):
        """
        Processing loop for the spans within an indexed section.

        Args:
            datasheet (dict): datasheet output document.
            sheetKey (string): worksheet key.
            wb (object): workbook object_
            sheetName (string): name of current worksheet
            section (dict): describes processing rules within the map file for a section.
            spans (list): list of columns to process in the current section
            fieldHeaders (list): field headers to use in writing datasheet
            indexFieldName (string): field name of the index.
            rowValues (object): a row of values from the worksheet
            map (dict): parser rules for the worksheet
        """
        try:
            if indexOn is not None:
                columnIndex = -1

                colsToIgnore = map.getColsToIgnore(sheetName)
                ignoreBlanksFlag = map.ignoreBlanks(sheetName)
                # header = map.getValue(spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD, indexOn)
                indexOnRow = section[spreadsheettypes.SPREADSHEET_MAP_FIELDHEADERS_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                indexOnCol = section[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD][0]
                headerCol = indexOnCol

                # dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, indexOn)
                indexFieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb, sheetName, indexOnRow, indexOnCol))

                if spans is not None and len(spans) > 0:
                    for s in spans:
                        if s in colsToIgnore:
                            columnIndex += 1
                            pass
                        else:
                            val = None
                            itemDict = None
                            groupBy = None
                            # groupByItemHeaderDictList = None
                            # groupByItemHeaders = None
                            # itemSubDict = None
                            itemSubNewDict = None
                            # deepKey = None
                            # newDict = None
                            # itemKeyName = None
                            columnIndex += 1

                            val = self.getColValueFromRow(rowValues, s, wb, rowNumber)

                            # if val is None:
                            #     val = spreadsheettypes.SPREADSHEET_MAP_EMPTY_VAL
                            if ignoreBlanksFlag is True and len(str(val).strip()) == 0:
                                pass
                            elif ((val is not None) and (len(str(val)) >= 0)):
                                # this is the value from the key value pair to look up in the generated datasheet

                                # if isinstance(val, str):
                                #     val = val.strip()

                                #     val = val.replace('\n', '')

                                indexValue = self.getColValueFromRow(rowValues, headerCol, wb, rowNumber)

                                # itemDict is the sub-document containing the key.

                                # check whether this is item is in a GroupBy.  If it is, get the sub-document and insert new value
                                groupBy = map.getGroups(section)
                                # groupByList = SpreadsheetMap.getGroupBySpansList(section)

                                # if groupBy is None:
                                #   itemDict = self.getDatasheetDict(datasheet, sheetKey, indexFieldName, indexValue)
                                # groupBy = None

                                # if SpreadsheetMap.columnInGroupByList(s, groupBy):

                                #         # Build the header list for the sub-document field names to be added
                                #         groupByItemHeaderDictList = SpreadsheetMap.getGroupByItemHeaderDictList(groupBy)
                                #         groupByItemHeaders = self.getFieldHeaderList(
                                #             wb, sheetName, groupByItemHeaderDictList)  # get a list of the item headers

                                #         # itemDict = self.getDatasheetDictGroupBy(datasheet, sheetKey, indexFieldName, indexValue, groupByItemHeaders[0])
                                #         itemDict = datasheet[sheetKey][rowNumTracker]

                                #         itemSubDict = itemDict.get(groupByItemHeaders[0])  # get from a dictionary

                                #         # see if there is a subgroup in the itemDict
                                #         if itemSubDict is not None:

                                #             # itemSubDict already exists so append the current key/value pair to the sub-document
                                #             # the document already exists, but when we pull this back, it is missing the top level padParameters key add it back in
                                #             itemSubDict = dict({groupByItemHeaders[0]: itemSubDict})

                                #             deepKey = ""
                                #             for g in groupByItemHeaders:
                                #                 deepKey = deepKey + g + '.'
                                #             deepKey = deepKey + fieldHeaders[columnIndex]

                                #         else:

                                #             # No sub dictionary found
                                #             # Create an empty dictionary with the top level groupby header
                                #             itemSubDict = dict({})

                                #             deepKey = ""
                                #             groupIndex = 0
                                #             for g in groupByItemHeaders:
                                #                 groupToValidate = groupBy[groupIndex]
                                #                 groupSpans = groupToValidate[spreadsheettypes.SPREADSHEET_MAP_SPANS_FIELD][spreadsheettypes.SPREADSHEET_MAP_COLUMNS_FIELD]
                                #                 # Only add the grouping for the spans section
                                #                 if s in groupSpans:
                                #                     deepKey = deepKey + g + '-'
                                #                 groupIndex += 1
                                #             deepKey = deepKey + fieldHeaders[columnIndex]

                                #             if len(groupByItemHeaders) > 0:
                                #                 for ih in groupByItemHeaders[1::]:

                                #                     newDict = dict({ih: {}})
                                #                     self.insertNestedDictionary(itemSubDict, newDict, groupByItemHeaders, groupByItemHeaders[0])

                                #         # self.set(itemSubDict, deepKey, val)

                                #         # save the sub-document back to the index
                                #         # itemKeyName = groupByItemHeaders[0]
                                #         itemDict[deepKey] = val

                                # else:

                                #     # Not a groupBy Element
                                #     itemDict = dict({fieldHeaders[columnIndex]: val})
                                if groupBy is not None:

                                    spanListInGroupBy = SpreadsheetMap.getListSpanInGroupBy(section)
                                    if s in spanListInGroupBy and SpreadsheetMap.getSubObjectFlag(groupBy, s):
                                        compoundNameList = fieldHeaders[columnIndex].split('-')
                                        baseHeader = compoundNameList[0]
                                        subHeader = compoundNameList[1]
                                        itemSubNewDict = dict({subHeader: val})
                                        itemDict = dict({baseHeader: itemSubNewDict})
                                    else:
                                        if val is not None and len(str(val)) >= 0:
                                            itemDict = dict({fieldHeaders[columnIndex]: val})
                                else:
                                    # this is not a groupBy element so just add it
                                    # New value will be inserted into the sub-document and the datasheet updated

                                    if val is not None and len(str(val)) >= 0:

                                        # ExceptionLogger.logDebug(__name__,"fieldHeaders=",fieldHeaders)
                                        # ExceptionLogger.logDebug(__name__,"columnIndex=",columnIndex)
                                        # ExceptionLogger.logDebug(__name__,"itemDict=",itemDict)
                                        itemDict = dict({fieldHeaders[columnIndex]: val})

                                if itemDict is not None:

                                    self.updateDatasheetDict(datasheet, sheetKey, indexFieldName, indexValue, itemDict, rowNumber, dataStartOffset, rowNumTracker, map, section, entry_index, sheetName)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def processSpanValuesForNonIndexedSection(self, datasheet, sheetKey, wb, sheetName, section, spans, fieldHeaders, rowNumber, rowValues, map):
        """
        Processing loop for the spans within a non-indexed section (no key row).

        Args:
            datasheet (dict): datasheet output document.
            sheetKey (string): worksheet key.
            wb (object): workbook object_
            sheetName (string): name of current worksheet
            section (dict): describes processing rules within the map file for a section.
            spans (list): list of columns to process in the current section
            fieldHeaders (list): field headers to use in writing datasheet
            indexFieldName (string): field name of the index.
            rowValues (object): a row of values from the worksheet
            map (dict): parser rules for the worksheet
        """
        try:

            columnIndex = -1

            # dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, section)
            # maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, section)
            itemSubDict = dict({})

            # ExceptionLogger.logDebug(__name__,"fieldHeaders=",fieldHeaders)

            if spans is not None and len(spans) > 0:
                for i in range(len(spans)):
                    s = spans[i]
                    val = None
                    groupBy = None
                    groupByItemHeaderDictList = None
                    groupByItemHeaders = None
                    deepKey = None
                    newDict = None
                    # itemKeyName = None
                    columnIndex += 1

                    val = self.getColValueFromRow(rowValues, s)

                    st = "Cell " + str(s) + str(rowNumber + 1) + ' value='
                    ExceptionLogger.logDebug(__name__, st, val)

                    if val is not None or len(str(val)) < 1:
                        ExceptionLogger.logDebug(__name__, "No value retrieved so will check and see if this is a merged cell..")
                        # No value retrieved so check if this is a merged cell.
                        if self.isMerged(wb, rowNumber + 1, s):
                            ExceptionLogger.logDebug(__name__, "This is a MERGED Cell, so we will attempt to get the value from the merged cell.")
                            val = self.getCellValueIfMerged(wb, rowNumber + 1, s)
                            ExceptionLogger.logDebug(__name__, "MERGED CELL,val=", val)

                    if ((val is not None) and (len(str(val)) > 0)):

                        if isinstance(val, str):
                            val = val.strip()

                            val = val.replace('\n', '')

                        # check whether this is item is in a GroupBy.  If it is, get the sub-document and insert new value
                        groupBy = map.getGroups(section)
                        # groupByList = SpreadsheetMap.getGroupBySpansList(section)

                        if groupBy is not None:
                            if SpreadsheetMap.columnInGroupByList(s, groupBy):

                                # Build the header list for the sub-document field names to be added
                                groupByItemHeaderDictList = SpreadsheetMap.getGroupByItemHeaderDictList(groupBy)
                                groupByItemHeaders = self.getFieldHeaderList(
                                    wb, sheetName, groupByItemHeaderDictList, map)  # get a list of the item headers

                                deepKey = ""
                                for g in groupByItemHeaders:
                                    deepKey = deepKey + g + '.'
                                deepKey = deepKey + fieldHeaders[columnIndex]

                                if len(groupByItemHeaders) > 0:
                                    for ih in groupByItemHeaders[1::]:

                                        newDict = dict({ih: {}})
                                        self.insertNestedDictionary(itemSubDict, newDict, groupByItemHeaders, groupByItemHeaders[0])

                                self.set(itemSubDict, deepKey, val)

                                # save the sub-document back to the index
                                # itemKeyName = groupByItemHeaders[0]
                                # ExceptionLogger.logDebug(__name__,"Adding itemSubDict[filedHeaders[columnIndex]]=",itemSubDict)
                                # itemSubDict[fieldHeaders[columnIndex]] = val
                            else:
                                # Not a groupBy Element
                                dbgStr = "not a GroupByElement:  fieldHeaders[columnIndex]=" + fieldHeaders[columnIndex] + ", val="
                                ExceptionLogger.logDebug(__name__, dbgStr, val)
                                itemSubDict[fieldHeaders[columnIndex]] = val
                        else:
                            # This is not a groupByElement
                            dbgStr = "GroupBy is None:  fieldHeaders[columnIndex]=" + fieldHeaders[columnIndex] + ", val="
                            ExceptionLogger.logDebug(__name__, dbgStr, val)
                            itemSubDict[fieldHeaders[columnIndex]] = val
                            # #this is not a groupBy element so just add it
                            # #New value will be inserted into the sub-document and the datasheet updated
                            # ExceptionLogger.logDebug(__name__,"itemSubDict[filedHeaders[columnIndex]]=",itemSubDict)
                            # ExceptionLogger.logDebug(__name__,"itemSubDict=",itemSubDict)
                            # ExceptionLogger.logDebug(__name__,"datasheet[sheetKey]=",datasheet[sheetKey])
                            # ExceptionLogger.logDebug(__name__,"itemSubDict=",itemSubDict)

                            # ExceptionLogger.logDebug(__name__,"datasheet[sheetKey]=",datasheet[sheetKey])
                    else:
                        # value is none
                        dbgStr = "val is NOT None so will see if " + s + str(rowNumber) + ' is a merged cell.  Val='
                        ExceptionLogger.logDebug(__name__, dbgStr, val)
                        # see if this is a merged cell
                        if self.isMerged(wb, rowNumber, s):
                            ExceptionLogger.logDebug(__name__, "MERGED CELL")

                if itemSubDict is not None and len(itemSubDict) > 0:
                    datasheet[sheetKey].append(itemSubDict)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def addSectionColumnsToDatasheet(self, datasheet, sheetKey, wb, sheetName, section, fieldHeaders, rowNum, rowValues, map, dataStartOffset, rowNumTracker, entry_index):
        """The main method that adds columns in a section to a datasheet.   This is one of the key functions in datasheet production from spreadsheets.

        Args:
            datasheet (dict): the output datasheet.
            sheetKey (string): the key name for the worksheet output in the datasheet.
            wb (object): workbook object.
            sheetName (string): worksheet name.
            section (dict): parsing rules for the current section
            fieldHeaders (list): contains the field headers (field names) for the column data.
            rowValues (object): a row object containing data to be written to the datasheet.
            map (dict): parsing rules for the workbook.
        """
        try:

            # columnIndex = -1
            # colLetter = ""
            indexOn = map.getIndexOn(section)
            spans = map.getSpansColumns(section)

            if indexOn is not None:

                if map.checkIndustryFormat():
                    # Write the row values for a section that is indexed in WG processing
                    self.processSpanValuesForIndexedSectionWG(datasheet, sheetKey, wb, sheetName, section,
                                                              spans, fieldHeaders, indexOn, rowNum, rowValues, map, dataStartOffset, rowNumTracker, entry_index)

                else:
                    # Write the row values for a section that is indexed
                    self.processSpanValuesForIndexedSection(datasheet, sheetKey, wb, sheetName, section,
                                                            spans, fieldHeaders, indexOn, rowNum, rowValues, map, dataStartOffset, rowNumTracker, entry_index)
            else:
                # Section is not indexed so process each row
                self.processSpanValuesForNonIndexedSection(datasheet, sheetKey, wb, sheetName, section,
                                                           spans, fieldHeaders, rowNum, rowValues, map)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def insertNestedDictionary(self, currentDict, newDict, keyList, name):
        """A recursive function that inserts a dictionary object at the lowest level within the referenced dictionary object.

        Args:
            currentDict (dict): current document.
            newDict (dict): dictionary item to insert at lowest/deepest level.
            keyList (list): a list of keys in the current dictionary.
            name (string): used by recursive function to traverse  dictionary.
        """
        try:
            key_in_dict = None
            for key in keyList:
                if key in currentDict:
                    key_in_dict = key
                break
            if key_in_dict is not None:
                self.insertNestedDictionary(currentDict[key_in_dict], newDict, keyList, name)
            else:
                currentDict[name] = newDict
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def depth(self, d):
        """Returns the depth of the referenced dictionary.

        Args:
            d (dict): dictionary to traverse.

        Returns:
            number: integer value indicating depth.
        """

        try:
            queue = deque([(id(d), d, 1)])
            memo = set()
            while queue:
                id_, o, level = queue.popleft()
                if id_ in memo:
                    continue
                memo.add(id_)
                if isinstance(o, dict):
                    queue += ((id(v), v, level + 1) for v in o.values())

            return level
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def set(self, d, key, value):
        """Sets a key/value in a dictionary

        Args:
            d (dict): a dictionary.
            key (string): specifies value of a key/value pair to set.
            value (object): the value to set.
        """

        # ExceptionLogger.logDebug(__name__,"Inside set",value)

        dd = d
        keys = key.split('.')
        latest = keys.pop()
        for k in keys:
            dd = dd.setdefault(k, {})
        dd.setdefault(latest, value)
        # ExceptionLogger.logDebug(__name__,"",value)

    def findDeepestKey(self, data):
        """Returns the key of the deepest sub-dictionary

        Args:
            data (dict): a dictionary to traverse.

        Returns:
            string: the key of the deepest dictionary item.
        """
        try:
            if not any([isinstance(data.get(k), dict) for k in data]):
                return data
            else:
                for dkey in data:
                    if isinstance(data.get(dkey), dict):
                        return self.findDeepestKey(data.get(dkey))
                    else:
                        continue
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def parseSectionWorkGroup(self, datasheet, sheetKey, wb, sheetName, section, map):
        """Main processing for an individual section to conform with WorkGroup Specifications.

        Args:
            datasheet (dict): the output datasheet.
            sheetKey (string): the key for the section in the datasheet.
            wb (object): a workbook object.
            sheetName (string): the name of the current worksheet.
            section (dict): parser rules for the current section.
            map (dict): parser rules for the workbook.

        Raises:
            ConfigurationError: returned if at least one section is not found.
        """
        try:
            dataStartRow = None
            maxRows = datasheetconstants.DATASHEET_DEFAULT_MAX_ROWS
            tableName = sheetKey

            if section is not None:

                fieldHeaders, fieldHeaderRow = self.getFieldHeadersForSection(wb, sheetName, section, map)
                indexOn = map.getIndexOn(section)
                if indexOn is not None:
                    maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, indexOn)
                    dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, indexOn)
                    header = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD]
                    if isinstance(header, list):
                        headerFieldName = ""
                        for header_element in header:
                            headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header_element)
                            headerCol = map.getValue(spreadsheettypes.SPREADSHEET_MAP_COL_FIELD, header_element)
                            headerFieldName_temp = self.getCellValue(wb, sheetName, headerRow, headerCol)
                            # Eliminate unwanted characters
                            headerFieldName_temp = self.format.delete_repeated_characters(headerFieldName_temp)
                            headerFieldName += f"{JsonDataSheet.generateValidJsonFieldName(headerFieldName_temp)}-"
                        # Eliminate last "-" character that it is not necessary
                        headerFieldName = headerFieldName[:-1]
                        # Save the header row to compare and validate if title is needed
                        headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header[0])
                        indexOnRow = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][0][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                        indexOnCol = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][0][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]
                    else:
                        # Just one object on the header section
                        headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header)
                        headerCol = map.getValue(spreadsheettypes.SPREADSHEET_MAP_COL_FIELD, header)
                        headerFieldName = self.getCellValue(wb, sheetName, headerRow, headerCol)
                        headerFieldName = self.format.delete_repeated_characters(headerFieldName)
                        headerFieldName = JsonDataSheet.generateValidJsonFieldName(headerFieldName)
                        indexOnRow = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                        indexOnCol = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]

                    self._indexOnRow = indexOnRow
                    self._indexOnCol = indexOnCol

                    # add the indexOn dictionary to the datasheet
                    self.serializeIndexWG(datasheet, tableName, wb, sheetName, section, map)

                    rowNumTracker = self.getRowNumberTracker(wb, map, section, datasheet, tableName, sheetName)
                    dataStartOffset = dataStartRow - 1
                    rowNum = dataStartOffset
                    rowsToIgnore = map.getRowsToIgnore(sheetName)

                else:
                    datasheet[tableName] = []
                    # IndexOn was not found so just process each row/column
                    dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, section)
                    if dataStartRow is None:
                        dataStartRow = datasheetconstants.DATASHEET_DEFAULT_DATASTART
                    maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, section)
                    if maxRows is None:
                        maxRows = datasheetconstants.DATASHEET_DEFAULT_MAX_ROWS
                    dataStartOffset = dataStartRow - 1
                    rowNum = dataStartOffset

                entry_index = 0
                entries_to_delete = []
                while rowNum < (maxRows + dataStartRow):
                    if indexOn is not None:
                        if rowNum + 1 in rowsToIgnore:
                            # Need to eliminate index entry from the datasheet
                            entries_to_delete.append(entry_index)
                            entry_index += 1
                            rowNum += 1
                            rowNumTracker += 1
                            continue

                    rowValues = self.getRow(wb, sheetName, rowNum)

                    if rowValues is not None:

                        # add the columns identified in the section configuration
                        if indexOn is not None:
                            self.addSectionColumnsToDatasheet(datasheet, tableName, wb, sheetName, section, fieldHeaders, rowNum, rowValues, map, dataStartOffset, rowNumTracker, entry_index)
                            rowNumTracker += 1
                        else:
                            self.addSectionColumnsToDatasheet(datasheet, tableName, wb, sheetName, section, fieldHeaders, rowNum, rowValues, map, dataStartOffset, rowNumTracker=None, entry_index=entry_index)
                    else:
                        break

                    rowNum += 1
                    entry_index += 1

            else:
                # did not find a section, raise an error
                s = t('Expected item not found')
                d = str(spreadsheettypes.SPREADSHEET_MAP_INDEX_ON_FIELD)
                raise ConfigurationError(s, d)
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def parseSection(self, datasheet, sheetKey, wb, sheetName, section, map):
        """Main processing for an individual section.

        Args:
            datasheet (dict): the output datasheet.
            sheetKey (string): the key for the section in the datasheet.
            wb (object): a workbook object.
            sheetName (string): the name of the current worksheet.
            section (dict): parser rules for the current section.
            map (dict): parser rules for the workbook.

        Raises:
            ConfigurationError: returned if at least one section is not found.
        """
        try:
            # orientation = spreadsheettypes.SPREADSHEET_MAP_ORIENTATION_FIELD
            # header = None
            # headerRow = None
            # headerCol = None
            dataStartRow = None
            # groupList = []
            maxRows = datasheetconstants.DATASHEET_DEFAULT_MAX_ROWS
            # headerFieldName = None
            tableName = sheetKey

            if section is not None:

                fieldHeaders, fieldHeaderRow = self.getFieldHeadersForSection(wb, sheetName, section, map)
                # ExceptionLogger.logDebug(__name__,"fieldHeaders:",fieldHeaders)

                # returns a list of the groupBy entries for the current section
                # groupList = self.parseGroups(section, map)  # get the groups in a section

                indexOn = map.getIndexOn(section)
                # ExceptionLogger.logDebug(__name__,"indexOn:",indexOn)

                # Check whether an indexOn was specified
                if indexOn is not None:
                    # indexFieldName = JsonDataSheet.generateValidJsonFieldName(self.getCellValue(wb, sheetName, indexOnRow, indexOnCol))
                    maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, indexOn)
                    # ExceptionLogger.logDebug(__name__,"indexFieldName=",indexFieldName)
                    dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, indexOn)

                    # Create table entry for the indicated header
                    # header = map.getValue(spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD, indexOn)
                    header = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD]
                    if isinstance(header, list):
                        # If it is a list, process all objects
                        headerFieldName = ""
                        for header_element in header:
                            headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header_element)
                            headerCol = map.getValue(spreadsheettypes.SPREADSHEET_MAP_COL_FIELD, header_element)
                            headerFieldName_temp = self.getCellValue(wb, sheetName, headerRow, headerCol)
                            # Eliminate unwanted characters
                            headerFieldName_temp = self.format.delete_repeated_characters(headerFieldName_temp)
                            headerFieldName += f"{JsonDataSheet.generateValidJsonFieldName(headerFieldName_temp)}-"
                        # Eliminate last "-" character that it is not necessary
                        headerFieldName = headerFieldName[:-1]
                        # Save the header row to compare and validate if title is needed
                        headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header[0])
                        indexOnRow = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][0][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                        indexOnCol = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][0][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]
                    else:
                        # Just one object on the header section
                        headerRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD, header)
                        headerCol = map.getValue(spreadsheettypes.SPREADSHEET_MAP_COL_FIELD, header)
                        headerFieldName = self.getCellValue(wb, sheetName, headerRow, headerCol)
                        headerFieldName = self.format.delete_repeated_characters(headerFieldName)
                        headerFieldName = JsonDataSheet.generateValidJsonFieldName(headerFieldName)
                        indexOnRow = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD]
                        indexOnCol = indexOn[spreadsheettypes.SPREADSHEET_MAP_HEADER_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD]

                    # only add the header value to the table name if it is not a table title
                    checkOnlyTableName = map.onlyTableName(sheetName)
                    if checkOnlyTableName:
                        tableName = headerFieldName
                        if sheetName not in datasheet:
                            datasheet[sheetName] = {}
                            datasheet[sheetName][tableName] = []
                        else:
                            datasheet[sheetName][tableName] = []
                    elif fieldHeaderRow != headerRow:
                        tableName = f"{sheetKey}-{headerFieldName}"
                        datasheet[tableName] = []
                    else:
                        tableName = f"{sheetKey}-table{self.tableCounter}"
                        datasheet[tableName] = []

                    # ExceptionLogger.logDebug(__name__, "Checking to see if we need to write the index")
                    # if len(datasheet[sheetKey])<1:
                    # if self._indexOnRow != indexOnRow or self._indexOnCol != indexOnCol:

                    self._indexOnRow = indexOnRow
                    self._indexOnCol = indexOnCol

                    # ExceptionLogger.logDebug(__name__,"++++++WRITING the Index For This Sheet",datasheet)

                    # add the indexOn dictionary to the datasheet
                    self.serializeIndex(datasheet, tableName, wb, sheetName, section, map)
                    # ExceptionLogger.logDebug(__name__,"++++++AFTER WRITING the Index For This Sheet",datasheet)

                    rowNumTracker = self.getRowNumberTracker(wb, map, section, datasheet, tableName, sheetName)
                    dataStartOffset = dataStartRow - 1
                    rowNum = dataStartOffset
                    rowsToIgnore = map.getRowsToIgnore(sheetName)
                    # dsRowIdx = -1

                else:
                    datasheet[tableName] = []
                    # IndexOn was not found so just process each row/column
                    dataStartRow = map.getValue(spreadsheettypes.SPREADSHEET_MAP_DATA_START_ROW_FIELD, section)
                    if dataStartRow is None:
                        dataStartRow = datasheetconstants.DATASHEET_DEFAULT_DATASTART
                    maxRows = map.getValue(spreadsheettypes.SPREADSHEET_MAP_MAX_ROWS_FIELD, section)
                    if maxRows is None:
                        maxRows = datasheetconstants.DATASHEET_DEFAULT_MAX_ROWS
                    dataStartOffset = dataStartRow - 1
                    rowNum = dataStartOffset
                    # ExceptionLogger.logDebug(__name__,"IndexOn not found so processing rows")

                entry_index = 0
                entries_to_delete = []
                while rowNum < (maxRows + dataStartRow):

                    # ExceptionLogger.logDebug(__name__,"++++++++++++++++++++++++++New Row.   Row Number=",rowNum)
                    if indexOn is not None:
                        if rowNum + 1 in rowsToIgnore:
                            # Need to eliminate index entry from the datasheet
                            entries_to_delete.append(entry_index)
                            entry_index += 1
                            rowNum += 1
                            rowNumTracker += 1
                            continue

                    rowValues = self.getRow(wb, sheetName, rowNum)

                    if rowValues is not None:

                        # add the columns identified in the section configuration
                        if indexOn is not None:
                            self.addSectionColumnsToDatasheet(datasheet, tableName, wb, sheetName, section, fieldHeaders, rowNum, rowValues, map, dataStartOffset, rowNumTracker, entry_index)
                            rowNumTracker += 1
                        else:
                            self.addSectionColumnsToDatasheet(datasheet, tableName, wb, sheetName, section, fieldHeaders, rowNum, rowValues, map, dataStartOffset, rowNumTracker=None, entry_index=entry_index)
                    else:
                        break

                    rowNum += 1
                    entry_index += 1

                    # ExceptionLogger.logDebug(__name__,"---------------------End of Row Number=",rowNum-1)
                if map.onlyTableName(sheetName):
                    key_elements = datasheet[sheetName][tableName]
                else:
                    key_elements = datasheet[tableName]
                # Reverse elements to delete
                entries_to_delete.sort(reverse=True)
                for element in entries_to_delete:
                    del key_elements[element]

            else:
                # did not find a section, raise an error
                s = t('Expected item not found')
                d = str(spreadsheettypes.SPREADSHEET_MAP_INDEX_ON_FIELD)
                raise ConfigurationError(s, d)

                # groups = map.getGroups(section)  # get the groups in a section

            # ExceptionLogger.logDebug(__name__,"datasheet=",datasheet)

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def getRowNumberTracker(self, wb, map, section, datasheet, sheetKey, sheetName):
        try:
            tagList = map.getIncludeTagsDictionaryList(section)
            if tagList is None:
                return 0
            elif tagList is not None and len(tagList) > 0:
                for tag in tagList:
                    fieldLabel = self.getCellValue(wb, sheetName,
                                                   tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                   tag[spreadsheettypes.SPREADSHEET_MAP_FIELDLABEL_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                    fieldValue = self.getCellValue(wb, sheetName,
                                                   tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_ROW_FIELD],
                                                   tag[spreadsheettypes.SPREADSHEET_MAP_FIELDVALUE_FIELD][spreadsheettypes.SPREADSHEET_MAP_COL_FIELD])
                    fieldLabel = JsonDataSheet.generateValidJsonFieldName(fieldLabel)
                    for i in range(len(datasheet[sheetKey])):
                        if fieldLabel in datasheet[sheetKey][i] and (datasheet[sheetKey][i][fieldLabel] == fieldValue):
                            rowNumTracker = i
                            return rowNumTracker
            return rowNumTracker
        except Exception:
            return 0

    def parseGroups(self, jsonObj, map):
        """Main processing method for a data grouping within a section.  This enables fields to be grouped under sub-documents.

        Args:
            jsonObj (dict): parser rules for the section.  These may contain optional groupBy elements.
            map (dict): parser rules for the worksheet.

        Returns:
            list: a list of groupBy elements.
        """
        parsed_groups = []
        try:
            groups = map.getGroups(jsonObj)

            if ((groups is not None) and (len(groups) > 0)):
                for group in groups:
                    parsed_groups.append(group)

            # debug message
            # ExceptionLogger.logDebug(__name__,"Groups:", groups)
            return parsed_groups

        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)

    def parseGroup(self, datasheet, wb, sheetName, group, map):
        """Main processing method for an individual grouping (groupBy).

        Args:
            datasheet (dict): the output datasheet.
            wb (object): the workbook.
            sheetName (string): the worksheet name.
            group (dict): a grouping rule.
            map (dict): the parser rules for the current workbook/worksheet.
        """
        try:
            pass
        except Exception as e:
            ExceptionLogger.logError(__name__, "", e)
